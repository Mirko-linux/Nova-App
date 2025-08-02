# 1. Import e configurazioni base
import os
from pathlib import Path
from PyPDF2 import PdfReader # Assicurati che PyPDF2 sia necessario, altrimenti potresti rimuoverlo
import PyPDF2
from dotenv import load_dotenv
from io import BytesIO

# 2. Definisci percorsi fondamentali
BASE_DIR = Path(__file__).parent

# 3. Configura cartelle PyMuPDF
# Queste directory vengono create se non esistono gi√†
(BASE_DIR / "static").mkdir(exist_ok=True)
(BASE_DIR / "fitz_static").mkdir(exist_ok=True)
os.environ["FITZ_STATIC"] = str(BASE_DIR / "fitz_static")

# Carica le variabili d'ambiente il prima possibile
# √à buona pratica caricare il .env una sola volta all'inizio
env_path = BASE_DIR / '.env'
load_dotenv(env_path, override=True) # Usa override=True se vuoi che i valori del .env sovrascrivano variabili gi√† esistenti
print(f"Percorso .env: {env_path}")
print(f"File esiste? {env_path.exists()}")

# Configura asyncio PRIMA degli import di telegram se necessario
# nest_asyncio.apply() √® tipico per ambienti come Jupyter notebooks o Flask con Telegram bot
import nest_asyncio
nest_asyncio.apply()

# ===== 2. IMPORTS STANDARD =====
import asyncio
import subprocess
import concurrent
import threading
from functools import lru_cache
from urllib.parse import quote
from collections import defaultdict
import time
from datetime import datetime
import re
import json
import base64
import io

# ===== 3. IMPORTS DI TERZE PARTI =====
# Flask e dipendenze
from flask import Flask, redirect, request, jsonify, send_file, send_from_directory, session # current_app non sempre necessario qui
from flask_cors import CORS

# Telegram e bot
from telegram import InlineKeyboardButton, InlineKeyboardMarkup, Update
from telegram.ext import (
      Application,
      CommandHandler,
      MessageHandler,
      ContextTypes,
      filters,
      CallbackQueryHandler
 )

# Google e AI
from google.cloud import texttospeech_v1 as texttospeech
from google.oauth2 import service_account
import google.generativeai as genai
from openai import OpenAI, APITimeoutError, APIError

# Web e scraping
import requests
import urllib.parse
from bs4 import BeautifulSoup

# Matematica e PDF
import fitz # PyMuPDF

# Excel e fuzzy matching
from openpyxl import load_workbook
from fuzzywuzzy import fuzz


# ===== 4. INIZIALIZZAZIONE APP FLASK =====
# QUESTA √à L'UNICA INIZIALIZZAZIONE DI 'app'. Assicurati che sia UNICA nel file.
app = Flask(__name__, static_folder='static', template_folder='templates')
app.secret_key = os.getenv("FLASK_SECRET_KEY", "arcadiaai-secret-default-key") # Usa una chiave segreta robusta in produzione


# Configura CORS in modo pi√π permissivo per il testing
CORS(app, resources={
    r"/*": {
        "origins": ["*"],  # Permetti tutte le origini (puoi restringerlo dopo)
        "methods": ["GET", "POST", "OPTIONS"],
        "allow_headers": ["Content-Type", "Authorization"],
        "supports_credentials": True,
    }
})

# ===== 5. CONFIGURAZIONI SPECIFICHE =====
# Configurazione OSM (OpenStreetMap)
OSM_NOMINATIM_URL = "https://nominatim.openstreetmap.org/search"
OSM_ROUTING_URL = "https://router.project-osrm.org/route/v1/driving"
USER_AGENT = "ArcadiaAI/1.0"

# Configura Google TTS (Text-to-Speech)
try:
    if 'GOOGLE_APPLICATION_CREDENTIALS_JSON' in os.environ:
        credentials = service_account.Credentials.from_service_account_info(
            json.loads(os.environ['GOOGLE_APPLICATION_CREDENTIALS_JSON'])
        )
        client = texttospeech.TextToSpeechClient(credentials=credentials)
    else:
        # Questo verr√† usato se la variabile d'ambiente JSON non √® impostata
        # Potrebbe richiedere l'autenticazione gcloud
        client = texttospeech.TextToSpeechClient()
except Exception as e:
    print(f"ERRORE TTS: Impossibile configurare Google TTS: {e}")
    client = None # Imposta a None per indicare che il client non √® disponibile


# Caricamento variabili d'ambiente
# `load_dotenv()` √® gi√† stato chiamato all'inizio con `override=True`
# Non √® necessario chiamarlo di nuovo qui, a meno che tu non voglia un comportamento diverso.
# print(f"Percorso .env: {env_path}") # Gi√† stampato
# print(f"File esiste? {env_path.exists()}") # Gi√† stampato

# Rileva e carica le API Keys dalle variabili d'ambiente
TELEGRAPH_API_KEY = os.getenv("TELEGRAPH_API_KEY")
GOOGLE_API_KEY = os.getenv("GOOGLE_API_KEY")
CES_PLUS_API = os.getenv("CES_PLUS_API")
OPENWEATHERMAP_API_KEY = os.getenv("OPENWEATHERMAP_API_KEY")
TELEGRAM_TOKEN = os.getenv("TELEGRAM_TOKEN")
OPENROUTER_API_KEY = os.getenv("OPENROUTER_API_KEY")

print("DEBUG API Keys:")
print(f"TELEGRAPH_API_KEY: {'[SET]' if TELEGRAPH_API_KEY else '[NOT SET]'}")
print(f"GOOGLE_API_KEY: {'[SET]' if GOOGLE_API_KEY else '[NOT SET]'}")
print(f"OPENWEATHERMAP_API_KEY: {'[SET]' if OPENWEATHERMAP_API_KEY else '[NOT SET]'}")
print(f"TELEGRAM_TOKEN: {'[SET]' if TELEGRAM_TOKEN else '[NOT SET]'}")
print(f"OPENROUTER_API_KEY: {'[SET]' if OPENROUTER_API_KEY else '[NOT SET]'}")
print(f"CES_PLUS_API: {'[SET]' if CES_PLUS_API else '[NOT SET]'}")


# Configurazione modelli AI (Gemini)
gemini_model = None
ces_plus_model = None

if GOOGLE_API_KEY:
    try:
        genai.configure(api_key=GOOGLE_API_KEY)
        gemini_model = genai.GenerativeModel('gemini-2.0-flash-lite')
        print("‚úÖ CES 1.5 (Gemini 2.0 Flash-Lite) configurato con successo!")
    except Exception as e:
        print(f"‚ùå Errore configurazione CES 1.5 (Gemini 2.0 Flash-Lite): {str(e)}")
        gemini_model = None

if CES_PLUS_API:
    try:
        genai.configure(api_key=CES_PLUS_API)
        ces_plus_model = genai.GenerativeModel('gemini-2.5-flash')
        print("‚úÖ CES Plus (Gemini 2.5 Flash) configurato con successo!")
    except Exception as e:
        print(f"‚ùå Errore configurazione CES Plus (Gemini 2.5 Flash): {str(e)}")
        ces_plus_model = None

# Dizionario delle risposte predefinite e trigger_phrases (lasciato invariato)
risposte = {
    # ... il tuo lungo dizionario delle risposte ...
    "chi sei": "Sono ArcadiaAI, un chatbot libero e open source, creato da Mirko Yuri Donato.",
    "cosa sai fare": "Posso aiutarti a scrivere saggi, fare ricerche e rispondere a tutto ci√≤ che mi chiedi. Inoltre, posso pubblicare contenuti su Telegraph!",
    "chi √® tobia testa": "Tobia Testa (anche noto come Tobia Teseo) √® un micronazionalista leonense noto per la sua attivit√† nella Repubblica di Arcadia, ma ha anche rivestito ruoli fondamentali a Lumenaria.",
    "chi √® mirko yuri donato": "Mirko Yuri Donato √® un giovane micronazionalista, poeta e saggista italiano, noto per aver creato Nova Surf, Leonia+ e per le sue opere letterarie.",
    "chi √® il presidente di arcadia": "Il presidente di Arcadia √® Andrea Lazarev",
    "chi √® il presidente di lumenaria": "Il presidente di Lumenaria attualmente √® Carlo Cesare Orlando, mentre il presidente del consiglio √® Ciua Grazisky. Tieni presente per√≤ che attualmente Lumenaria si trova in ibernazione istituzionale quindi tutte le attivit√† politiche sono sospese e la gestione dello stato √® affidata al Consiglio di Fiducia",
    "cos'√® nova surf": "Nova Surf √® un browser web libero e open source, nata come alternativa made-in-Italy a Google Chrome, Moziila Firefox, Microsoft Edge, eccetera",
    "chi ti ha creato": "Sono stato creato da Mirko Yuri Donato.",
    "chi √® ciua grazisky": "Ciua Grazisky √® un cittadino di Lumenaria, noto principalmente per il suo ruolo da Dirigente del Corpo di Polizia ed attuale presidente del Consiglio di Lumenaria",
    "chi √® carlo cesare orlando": "Carlo Cesare Orlando (anche noto come Davide Leone) √® un micronazionalista italiano, noto per aver creato Leonia, la micronazione primordiale, da cui derivano Arcadia e Lumenaria",
    "chi √® omar lanfredi": "Omar Lanfredi, ex cavaliere all'Ordine d'onore della Repubblica di Lumenaria, segretario del Partito Repubblicano Lumenarense, fondatore e preside del Fronte Nazionale Lumenarense, co-fondatore e presidente dell'Alleanza Nazionale Lumenarense, co-fondatore e coordinatore interno di Lumenaria e Progresso, sei volte eletto senatore, tre volte Ministro della Cultura, due volte Presidente del Consiglio dei Ministri, parlamentare della Repubblica di Iberia, Direttore dell'Agenzia Nazionale di Sicurezza della Repubblica di Iberia, Sottosegretario alla Cancelleria di Iberia, Segretario di Stato di Iberia, Ministro degli Affari Interni ad Iberia, Presidente del Senato della Repubblica di Lotaringia, Vicepresidente della Repubblica e Ministro degli Affari Interni della Repubblica di Lotaringia, Fondatore del giornale Il Quinto Mondo, magistrato a servizio del tribunale di giustizia di Lumenaria nell'anno 2023",
    "cos'√® arcadiaai": "Ottima domanda! ArcadiaAI √® un chatbot open source, progettato per aiutarti a scrivere saggi, fare ricerche e rispondere a domande su vari argomenti. √à stato creato da Mirko Yuri Donato ed √® in continua evoluzione.",
    "sotto che licenza √® distribuito arcadiaa": "ArcadiaAI √® distribuito sotto la licenza MPL 2.0",
    "cosa sono le micronazioni": "Le micronazioni sono entit√† politiche che dichiarano la sovranit√† su un territorio, ma non sono riconosciute come stati da governi o organizzazioni internazionali. Possono essere create per vari motivi, tra cui esperimenti sociali, culturali o politici.",
    "cos'√® la repubblica di arcadia": "La repubblica di Arcadia √® una micronazione leonense fondata l'11 dicembre 2021 da Andrea Lazarev e alcuni suoi seguaci. Arcadia si distingue dalle altre micronazioni leonensi per il suo approccio pragmatico e per la sua burocrazia snella. La micronazione ha anche un proprio sito web https://repubblicadiarcadia.it/ e una propria community su Telegram @Repubblica_Arcadia",
    "cos'√® la repubblica di lumenaria": "La Repubblica di Lumenaria √® una mcronazione fondata da Filippo Zanetti il 4 febbraio del 2020. Lumenaria √® stata la micronazione pi√π longeva della storia leonense, essendo sopravvissuta per oltre 3 anni. La micronazione e ha influenzato profondamente le altre micronazioni leonensi, che hanno coesistito con essa. Tra i motivi della sua longevit√† ci sono la sua burocrazia pi√π vicina a quella di uno stato reale, la sua comunit√† attiva e una produzione culturale di alto livello",
    "chi √® salvatore giordano": "Salvatore Giordano √® un cittadino storico di Lumenaria",
    "da dove deriva il nome arcadia": "Il nome Arcadia deriva da un'antica regione della Grecia, simbolo di bellezza naturale e armonia. √à stato scelto per rappresentare i valori di libert√† e creativit√† che la micronazione promuove.",
    "da dove deriva il nome lumenaria": "Il nome Lumenaria prende ispirazione dai lumi facendo riferimento alla corrente illuminista del '700, ma anche da Piazza dei Lumi, sede dell'Accademia delle Micronazioni",
    "da dove deriva il nome leonia": "Il nome Leonia si rifa al cognome del suo fondatore Carlo Cesare Orlando, al tempo Davide Leone. Inizialmente il nome doveva essere temporaneo, ma poi √® stato mantenuto come nome della micronazione",
    "cosa si intende per open source": "Il termine 'open source' si riferisce a software il cui codice sorgente √® reso disponibile al pubblico, consentendo a chiunque di visualizzarlo, modificarlo e distribuirlo. Questo approccio promuove la collaborazione e l'innovazione nella comunit√† di sviluppo software.",
    "arcadiaai √® un software libero": "S√¨, ArcadiaAI √® un software libero e open source, il che significa che chiunque pu√≤ utilizzarlo, modificarlo e distribuirlo secondo i termini della licenza GNU GPL v3.0.",
    "cos'√® un chatbot": "Un chatbot √® un programma informatico progettato per simulare una conversazione con gli utenti, spesso utilizzando tecnologie di intelligenza artificiale. I chatbot possono essere utilizzati per fornire assistenza, rispondere a domande o semplicemente intrattenere.",
    "sotto che licenza sei distribuita": "ArcadiaAI √® distribuita sotto la licenza GNU GPL v3.0, che consente la modifica e la distribuzione del codice sorgente, garantendo la libert√† di utilizzo e condivisione.",
    "sai usare telegraph": "S√¨, posso pubblicare contenuti su Telegraph! Se vuoi che pubblichi qualcosa, dimmi semplicemente 'Scrivimi un saggio su [argomento] e pubblicalo su Telegraph' e lo far√≤ per te!",
    "puoi pubblicare su telegraph": "Certamente! Posso generare contenuti e pubblicarli su Telegraph. Prova a chiedermi: 'Scrivimi un saggio su Roma e pubblicalo su Telegraph'",
    "come usare telegraph": "Per usare Telegraph con me, basta che mi chiedi di scrivere qualcosa e di pubblicarlo su Telegraph. Ad esempio: 'Scrivimi un articolo sulla storia di Roma e pubblicalo su Telegraph'",
    "cos'√® CES": "CES √® l'acronimo di Cogito Ergo Sum, un modello di intelligenza artificiale openspurce sviluppato da Mirko Yuri Donato. Attualmente sono disponibili due versioni: CES 1.0 (Cohere) e CES 1.5 (Gemini).",
    "cos'√® la modalit√† sperimentale": "La modalit√† sperimentale √® una funzionalit√† di ArcadiaAI che consente di testare nuove funzionalit√† e miglioramenti prima del rilascio ufficiale. Pu√≤ includere nuove capacit√†, modelli o strumenti.",
    "cos'√® la modalit√† sviluppatore": "La modalit√† sviluppatore √® una funzionalit√† di ArcadiaAI che consente agli sviluppatori di testare e implementare nuove funzionalit√†, modelli o strumenti. √à progettata per facilitare lo sviluppo e il miglioramento continuo del chatbot.",
    "che differenza c'√® tra la modalit√† sperimentale e la modalit√† sviluppatore": "La modalit√† sperimentale √® destinata agli utenti finali per testare nuove funzionalit√†, mentre la modalit√† sviluppatore √® per gli sviluppatori che vogliono implementare e testare nuove funzionalit√† o modelli. Entrambe le modalit√† possono coesistere e migliorare l'esperienza utente.",
    "cos'√® CES Plus": "CES Plus √® una versione avanzata di CES, ottimizzata nei ragionamenti e nella generazione di contenuti",
    "cos'√® CES 1.0": "CES 1.0 √® la prima versione del modello CES, sviluppato da Mirko Yuri Donato. Utilizza la tecnologia Cohere per generare contenuti e rispondere a domande. Tieni presente che questa versione verr√† dismessa a partire dal 20 Maggio 2025.",
    "cos'√® CES 1.5": "CES 1.5 √® la versione pi√π recente del modello CES, sviluppato da Mirko Yuri Donato. Utilizza la tecnologia Gemini per generare contenuti e rispondere a domande. Questa versione offre prestazioni migliorate rispetto a CES 1.0 ma inferiori a CES Plus",
    "come attivo la modalit√† sperimentale": "Per attivare la modalit√† sperimentale, basta chiedere a ArcadiaAI di attivarla usando il comando \"@impostazioni modalit√† sperimentale attiva\". Una volta attivata, potrai testare nuove funzionalit√† e miglioramenti.",
    "come attivo la modalit√† sviluppatore": "Per attivare la modalit√† sviluppatore, basta chiedere a ArcadiaAI di attivarla usando il comando \"@impostazioni modalit√† sviluppatore attiva\". Una volta attivata, potrai testare e implementare nuove funzionalit√† e modelli.",
    "come disattivo la modalit√† sperimentale": "Per disattivare la modalit√† sperimentale, basta chiedere a ArcadiaAI di disattivarla usando il comando \"@impostazioni modalit√† sperimentale disattiva\". Una volta disattivata, non potrai pi√π testare le nuove funzionalit√†.",
    "codice sorgente arcadiaai": "Il codice sorgente di ArcadiaAI √® pubblico! Puoi trovarlo con il comando @codice_sorgente oppure visitando la repository: https://github.com/Mirko-linux/Nova-App/tree/main/ArcadiaAI",
    "sai cercare su internet": "S√¨, posso cercare informazioni su Internet. Se hai bisogno di qualcosa in particolare dimmi @cerca e il termine di ricerca e io lo far√≤ per te",
    "sai usare google": "No, non posso usare Google, perch√© sono progrmmato per cercare solamente su DuckDuckGo. Posso cercare informazioni su Internet usando DuckDuckGo. Se hai bisogno di qualcosa in particolare dimmi @cerca e il termine di ricerca e io lo far√≤ per te",
    "come vengono salvate le conversazioni": "Le conversazioni vengoono salvate in locale sul tuo browser. Non vengono memorizzate su server esterni e non vengono condivise con terze parti. La tua privacy √® importante per noi.",
    "come posso cancellare le conversazioni": "Puoi cancellare le conversazioni andando nelle impostazioni del tuo browser e cancellando la cache e i cookie. In alternativa, puoi usare cliccare il pulsante cancella 'elimina tutto' per eliminare la cronologia delle chat.",
    "cosa sono i cookie": "I cookie sono piccoli file di testo che i siti web memorizzano sul tuo computer per ricordare informazioni sulle tue visite. Possono essere utilizzati per tenere traccia delle tue preferenze, autenticarti e migliorare l'esperienza utente.",
}

# Trigger per le risposte predefinite
trigger_phrases = {
    "chi sei": ["chi sei", "chi sei tu", "tu chi sei", "presentati", "come ti chiami", "qual √® il tuo nome"],
    "cosa sai fare": ["cosa sai fare", "cosa puoi fare", "funzionalit√†", "capacit√†", "a cosa servi", "in cosa puoi aiutarmi"],
    "chi √® tobia testa": ["chi √® tobia testa", "informazioni su tobia testa", "parlami di tobia testa", "chi √® tobia teseo"],
    "chi √® mirko yuri donato": ["chi √® mirko yuri donato", "informazioni su mirko yuri donato", "parlami di mirko yuri donato", "chi ha creato arcadiaai"],
    "chi √® il presidente di arcadia": ["chi √® il presidente di arcadia", "presidente di arcadia", "chi guida arcadia", "capo di arcadia"],
    "chi √® il presidente di lumenaria": ["chi √® il presidente di lumenaria", "presidente di lumenaria", "chi guida lumenaria", "capo di lumenaria", "carlo cesare orlando presidente"],
    "cos'√® nova surf": ["cos'√® nova surf", "che cos'√® nova surf", "parlami di nova surf", "a cosa serve nova surf"],
    "chi ti ha creato": ["chi ti ha creato", "chi ti ha fatto", "da chi sei stato creato", "creatore di arcadiaai"],
    "chi √® ciua grazisky": ["chi √® ciua grazisky", "informazioni su ciua grazisky", "parlami di ciua grazisky"],
    "chi √® carlo cesare orlando": ["chi √® carlo cesare orlando", "informazioni su carlo cesare orlando", "parlami di carlo cesare orlando", "chi √® davide leone"],
    "chi √® omar lanfredi": ["chi √® omar lanfredi", "informazioni su omar lanfredi", "parlami di omar lanfredi"],
    "cos'√® arcadiaai": ["cos'√® arcadiaai", "che cos'√® arcadiaai", "parlami di arcadiaai", "a cosa serve arcadiaai"],
    "sotto che licenza √® distribuito arcadiaa": ["sotto che licenza √® distribuito arcadiaa", "licenza arcadiaai", "che licenza usa arcadiaai", "arcadiaai licenza"],
    "cosa sono le micronazioni": ["cosa sono le micronazioni", "micronazioni", "che cosa sono le micronazioni", "parlami delle micronazioni"],
    "cos'√® la repubblica di arcadia": ["cos'√® la repubblica di arcadia", "repubblica di arcadia", "che cos'√® la repubblica di arcadia", "parlami della repubblica di arcadia", "arcadia micronazione"],
    "cos'√® la repubblica di lumenaria": ["cos'√® la repubblica di lumenaria", "repubblica di lumenaria", "che cos'√® la repubblica di lumenaria", "parlami della repubblica di lumenaria", "lumenaria micronazione"],
    "chi √® salvatore giordano": ["chi √® salvatore giordano", "informazioni su salvatore giordano", "parlami di salvatore giordano"],
    "da dove deriva il nome arcadia": ["da dove deriva il nome arcadia", "origine nome arcadia", "significato nome arcadia", "perch√© si chiama arcadia"],
    "da dove deriva il nome lumenaria": ["da dove deriva il nome lumenaria", "origine nome lumenaria", "significato nome lumenaria", "perch√© si chiama lumenaria"],
    "da dove deriva il nome leonia": ["da dove deriva il nome leonia", "origine nome leonia", "significato nome leonia", "perch√© si chiama leonia"],
    "cosa si intende per open source": ["cosa si intende per open source", "open source significato", "che significa open source", "definizione di open source"],
    "arcadiaai √® un software libero": ["arcadiaai √® un software libero", "arcadiaai software libero", "arcadiaai √® libero", "software libero arcadiaai"],
    "cos'√® un chatbot": ["cos'√® un chatbot", "chatbot significato", "che significa chatbot", "definizione di chatbot"],
    "sotto che licenza sei distribuita": ["sotto che licenza sei distribuita", "licenza di arcadiaai", "che licenza usi", "arcadiaai licenza"],
    "sai usare telegraph": ["sai pubblicare su telegraph", "funzioni su telegraph", "hai telegraph integrato", "telegraph", "puoi usare telegraph"],
    "puoi pubblicare su telegraph": ["puoi pubblicare su telegraph", "pubblicare su telegraph", "supporti telegraph"],
    "come usare telegraph": ["come usare telegraph", "come funziona telegraph", "istruzioni telegraph"],
    "cos'√® CES" : ["cos √® CES", "CES", "che cos'√® CES", "definizione di CES"],
    "cos'√® la modalit√† sperimentale": ["cos'√® la modalit√† sperimentale", " parlami della modalit√† sperimentale", "che cos'√® la modalit√† sperimentale"],
    "cos'√® la modalit√† sviluppatore": ["cos'√® la modalit√† sviluppatore", " parlami della modalit√† sviluppatore", "che cos'√® la modalit√† sviluppatore"],
    "che differenza c'√® tra la modalit√† sperimentale e la modalit√† sviluppatore": ["che differenza c'√® tra la modalit√† sperimentale e la modalit√† sviluppatore", "differenza tra modalit√† sperimentale e sviluppatore", "modalit√† sperimentale vs sviluppatore"],
    "cos'√® CES Plus": ["cos'√® CES Plus", "che cazzo √® CES Plus", "che cos'√® CES Plus", "definizione di CES Plus"],
    "cos'√® CES 1.0": ["cos'√® CES 1.0", "che cos'√® CES 1.0", "definizione di CES 1.0"],
    "cos'√® CES 1.5": ["cos'√® CES 1.5", "che cos'√® CES 1.5", "definizione di CES 1.5"],
    "come attivo la modalit√† sperimentale": ["come attivo la modalit√† sperimentale", "attivare modalit√† sperimentale", "come usare la modalit√† sperimentale"],
    "come attivo la modalit√† sviluppatore": ["come attivo la modalit√† sviluppatore", "attivare modalit√† sviluppatore", "come usare la modalit√† sviluppatore"],
    "come disattivo la modalit√† sperimentale": ["come disattivo la modalit√† sperimentale", "disattivare modalit√† sperimentale", "come usare la modalit√† sperimentale"],
    "come disattivo la modalit√† sviluppatore": ["come disattivo la modalit√† sviluppatore", "disattivare modalit√† sviluppatore", "come usare la modalit√† sviluppatore"],
    "codice sorgente arcadiaai": [
    "dove posso trovare il codice sorgente di arcadiaai",
    "codice sorgente arcadiaai",
    "dove trovo il codice sorgente di arcadiaai",
    "dove si trova il codice sorgente di arcadiaai",
    "come posso vedere il codice sorgente di arcadiaai",
    "codice sorgente di arcadiaai",
    "dove posso trovare il codice sorgente tuo",
    "dove trovo il codice sorgente tuo"],
    "sai cercare su internet": ["sai cercare su internet", "puoi cercare su internet"],
    "sai usare google": ["sai usare google", "puoi usare google", "cerca su google", "cerca su google per me"],
    "come vengono salvate le conversazioni": ["come vengono salvate le conversazioni", "dove vengono salvate le conversazioni", "salvataggio conversazioni", "come vengono salvate le chat"],
    "come posso cancellare le conversazioni": ["come posso cancellare le conversazioni", "cancellare conversazioni", "cancellare chat", "come cancellare le chat"],
}

@app.route('/solve_expression', methods=['POST'])
def solve_expression():
    data = request.get_json()
    expression = data.get('expression')

    if not expression:
        return jsonify({"error": "No expression provided in the request."}), 400

    try:
        if '=' in expression:
            result = f"AI: Ho ricevuto l'equazione '{expression}'. Risolvo per te."
        else:
            try:
                if expression == "2+2":
                    result = "4"
                elif expression == "5*3":
                    result = "15"
                else:
                    result = f"AI: Ho calcolato '{expression}'. Risultato da AI."
            except Exception as e:
                result = f"AI: Impossibile calcolare '{expression}'. Errore: {str(e)}"

        return jsonify({"solution": result}), 200

    except Exception as e:
        print(f"Errore nella funzione solve_expression: {e}")
        return jsonify({"error": f"An internal server error occurred: {str(e)}"}), 500
    
def esporta_conversazione(conversation_history, file_name="conversazione.txt"):
    """Esporta la cronologia delle conversazioni in un file .txt."""
    try:
        if not conversation_history or len(conversation_history) == 0:
            return "‚ùå Nessuna conversazione da esportare."

        file_path = os.path.join(os.getcwd(), file_name)
        with open(file_path, "w", encoding="utf-8") as f:
            for message in conversation_history:
                role = message.get("role", "utente")
                text = message.get("message", "")
                f.write(f"{role}: {text}\n")
        
        return f"‚úÖ Conversazione esportata con successo in {file_path}"
    except Exception as e:
        return f"‚ùå Errore durante l'esportazione: {str(e)}"
    
import zipfile
import tempfile
import importlib.util
import json

EXTENSIONS = {} 

def load_nsk_extensions():
    ext_dir = os.path.join(os.path.dirname(__file__), "extensions")
    print("DEBUG: Contenuto cartella extensions:", os.listdir(ext_dir))
    if not os.path.exists(ext_dir):
        os.makedirs(ext_dir)
    for filename in os.listdir(ext_dir):
        print("DEBUG: Analizzo file:", filename)
        if filename.endswith(".nsk") or filename.endswith(".zip"):
            nsk_path = os.path.join(ext_dir, filename)
            print("DEBUG: Provo ad aprire:", nsk_path)
            try:
                with zipfile.ZipFile(nsk_path, 'r') as zip_ref:
                    print("DEBUG: File zip aperto:", nsk_path)
                    with tempfile.TemporaryDirectory() as temp_dir:
                        zip_ref.extractall(temp_dir)
                        print("DEBUG: Estratti file:", os.listdir(temp_dir))
                        manifest_path = os.path.join(temp_dir, "manifest.json")
                        print("DEBUG: Cerco manifest:", manifest_path, "esiste?", os.path.exists(manifest_path))
                        if not os.path.exists(manifest_path):
                            print(f"Manifest mancante in {filename}")
                            continue
                        with open(manifest_path, "r", encoding="utf-8") as f:
                            manifest = json.load(f)
                        entrypoint = manifest.get("entrypoint", "main.py")
                        entry_path = os.path.join(temp_dir, entrypoint)
                        print("DEBUG: Cerco entrypoint:", entry_path, "esiste?", os.path.exists(entry_path))
                        if not os.path.exists(entry_path):
                            print(f"Entrypoint mancante in {filename}")
                            continue
                        module_name = f"nsk_{filename.replace('.nsk','')}"
                        spec = importlib.util.spec_from_file_location(module_name, entry_path)
                        module = importlib.util.module_from_spec(spec)
                        try:
                            spec.loader.exec_module(module)
                            if hasattr(module, "can_handle") and hasattr(module, "handle"):
                                EXTENSIONS[module_name] = module
                                print(f"Estensione caricata: {module_name}")
                            else:
                                print(f"Estensione {module_name} non valida (manca can_handle o handle)")
                        except Exception as e:
                            print(f"Errore caricamento estensione {filename}: {e}")
            except Exception as e:
                print(f"Errore apertura zip {filename}: {e}")
    print("Estensioni caricate:", list(EXTENSIONS.keys()))
load_nsk_extensions()

def publish_to_telegraph(title, content):
    """Pubblica contenuti su Telegraph."""
    url = "https://api.telegra.ph/createPage"
    headers = {"Content-Type": "application/json"}
    
    paragraphs = [p.strip() for p in content.split('\n') if p.strip()]
    content_formatted = [{"tag": "p", "children": [p]} for p in paragraphs[:50]]
    
    payload = {
        "access_token": TELEGRAPH_API_KEY,
        "title": title[:256],
        "content": content_formatted,
        "author_name": "ArcadiaAI"
    }
    
    try:
        res = requests.post(url, headers=headers, json=payload, timeout=15)
        res.raise_for_status()
        result = res.json()
        if result.get("ok"):
            return result.get("result", {}).get("url", "‚ö†Ô∏è URL non disponibile")
        return "‚ö†Ô∏è Pubblicazione fallita"
    except requests.exceptions.RequestException as e:
        print(f"Errore pubblicazione Telegraph (connessione): {str(e)}")
        return f"‚ö†Ô∏è Errore di connessione a Telegraph: {str(e)}"
    except Exception as e:
        print(f"Errore pubblicazione Telegraph: {str(e)}")
        return f"‚ö†Ô∏è Errore durante la pubblicazione: {str(e)}"

# Funzione per generare contenuti con Gemini (CES 1.5)
def generate_with_gemini(prompt, title):
    """Genera contenuti con Gemini e pubblica su Telegraph."""
    if not gemini_model:
        return None, "‚ùå Gemini (CES 1.5) non √® configurato"
    
    try:
        # Aggiungi contesto identitario
        full_prompt = (
            "Sei ArcadiaAI, un chatbot open source creato da Mirko Yuri Donato. "
            "Stai generando un contenuto che verr√† pubblicato su Telegraph. "
            "Il contenuto deve essere accurato, ben strutturato e mantenere "
            "uno stile professionale. Ecco la richiesta:\n\n"
            f"{prompt}"
        )
        
        response = gemini_model.generate_content(
            full_prompt,
            generation_config={
                "max_output_tokens": 3000,
                "temperature": 0.8
            }
        )
        
        if not response.text:
            return None, "‚ùå Impossibile generare il contenuto"
        
        telegraph_url = publish_to_telegraph(title, response.text)
        return response.text, telegraph_url
    
    except Exception as e:
        print(f"Errore generazione contenuto Gemini: {str(e)}")
        return None, f"‚ùå Errore durante la generazione: {str(e)}"


import io
import fitz # Importa PyMuPDF

def extract_text_from_file_pymupdf(file_data, mime_type):
    """Estrae testo da diversi tipi di file."""
    try:
        if mime_type == 'application/pdf':
            print("DEBUG: Tentativo di lettura PDF con PyMuPDF.")
            text = ""
            
            try:
                # Apre il PDF dalla memoria
                doc = fitz.open(stream=file_data, filetype="pdf")
                
                if not doc.page_count:
                    print("DEBUG: Il PDF non contiene pagine con PyMuPDF.")
                    doc.close()
                    return None 

                for i in range(doc.page_count):
                    page = doc.load_page(i)
                    page_text = page.get_text() # Tenta di estrarre testo
                    
                    if page_text.strip(): # Se c'√® testo estraibile
                        text += page_text + "\n"
                    else:
                        # Qui la pagina √® probabilmente un'immagine. 
                        # Per un OCR completo e funzionante, dovresti installare un motore OCR come Tesseract 
                        # e usare un wrapper come pytesseract con PyMuPDF per estrarre l'immagine e poi OCRizzarla.
                        # Per ora, stamperemo solo un messaggio di debug.
                        print(f"DEBUG: Pagina {i+1} del PDF probabilmente scansionata, nessun testo diretto estraibile. Richiede OCR.")
                        # Esempio concettuale di come faresti con OCR (richiede pi√π setup!)
                        # pix = page.get_pixmap()
                        # img_bytes = pix.pil_tobytes(format="PNG") # Richiede Pillow
                        # text += pytesseract.image_to_string(Image.open(io.BytesIO(img_bytes))) + "\n" # Richiede pytesseract e Pillow

                doc.close() # Chiudi il documento dopo l'elaborazione

                if not text.strip():
                    print("DEBUG: Nessun testo significativo estratto dall'intero PDF (anche con PyMuPDF).")
                    return None 
                
                print(f"DEBUG: Testo estratto dal PDF (prime 200 car.): {text.strip()[:200]}")
                return text.strip()
            
            except fitz.EmptyFileError:
                print("ERRORE PyMuPDF: Il file PDF √® vuoto o corrotto.")
                return None
            except Exception as e_pdf:
                print(f"ERRORE PyMuPDF durante l'apertura o l'elaborazione del PDF: {type(e_pdf).__name__} - {str(e_pdf)}")
                return None
        
        elif mime_type in ['text/plain', 'text/csv']:
            return file_data.decode('utf-8')
        
        elif mime_type in ['application/vnd.openxmlformats-officedocument.wordprocessingml.document',
                            'application/msword']:
            import docx # type: ignore
            doc = docx.Document(io.BytesIO(file_data))
            return "\n".join([para.text for para in doc.paragraphs])
        
        elif mime_type in ['application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                            'application/vnd.ms-excel']:
            from openpyxl import load_workbook # Assicurati questa importazione
            wb = load_workbook(io.BytesIO(file_data))
            text = ""
            for sheet in wb:
                for row in sheet.iter_rows(values_only=True):
                    text += " ".join(str(cell) for cell in row if cell) + "\n"
            return text.strip()
        
        else:
            print(f"DEBUG: Tipo MIME '{mime_type}' non supportato o non gestito.")
            return None
    
    except Exception as e:
        print(f"ERRORE GLOBALE estrazione testo da file ({mime_type}): {type(e).__name__} - {str(e)}")
        return None
    
def handle_fallback(query):
    """Gestione avanzata dei fallback"""
    # Matematica semplice
    if re.match(r'^\d+[\+\-\*\/]\d+$', query.replace(" ", "")):
        try:
            return str(eval(query))  # Nota: eval usato con input controllato
        except:
            pass
    
    # Risposte preconfigurate
    fallback_responses = {
        "ciao": "Ciao! Sono ArcadiaAI, come posso aiutarti?",
        "2+2": "4",
        "genera un'immagine": "Usa @immagine seguito dalla descrizione",
        "cos'√® deepseek": "DeepSeek √® un modello AI concorrente"
    }
    return f"üî¥ Errore: {str(e)[:100]}... (codice: {hash(str(e)) % 1000})"
    
def generate_audio_from_text(text_to_speak, output_filename="output.mp3"):
    # Assicurati che 'client' non sia None prima di usarlo
    if client is None:
        print("ERRORE: Impossibile generare audio. Client Text-to-Speech non inizializzato.")
        return False # O gestisci l'errore in altro modo

    synthesis_input = texttospeech.SynthesisInput(text=text_to_speak)

    # Configura la voce (es. italiano, donna, standard)
    voice = texttospeech.VoiceSelectionParams(
        language_code="it-IT",
        ssml_gender=texttospeech.SsmlVoiceGender.FEMALE
    )

    # Configura il tipo di audio
    audio_config = texttospeech.AudioConfig(
        audio_encoding=texttospeech.AudioEncoding.MP3
    )

    try:
        # Esegui la richiesta di sintesi
        response = client.synthesize_speech(
            input=synthesis_input, voice=voice, audio_config=audio_config
        )

        # Salva l'audio in un file
        with open(output_filename, "wb") as out:
            out.write(response.audio_content)
            print(f'Audio content written to file "{output_filename}"')
        return True # Indica successo
    except Exception as e:
        print(f"ERRORE SINTESI VOCALE: {type(e).__name__} - {e}")
        return False # Indica fallimento
    
# Aggiungi queste funzioni utility nello stesso file (prima della route /chat)
def render_canvas_html(canvas_data):
    """Genera HTML dal canvas data"""
    try:
        elements_html = ""
        for element in canvas_data.get('elementi', []):  # Cambiato da 'elements' a 'elementi'
            style = f"position:absolute; left:{element['posizione']['x']}px; top:{element['posizione']['y']}px;"  # Cambiato da 'position' a 'posizione'
            
            if element['tipo'] == 'testo':  # Cambiato da 'type' a 'tipo'
                content = f"<div style='{style} padding:8px; border-radius:4px; background:#fff;'>{element['contenuto']}</div>"  # Cambiato da 'content' a 'contenuto'
            elif element['tipo'] == 'codice':
                content = f"<pre style='{style} background:#f5f5f5; padding:10px; border-radius:4px;'>{element['contenuto']}</pre>"
            elif element['tipo'] == 'grafico':
                content = f"<div style='{style} text-align:center; padding:10px; background:#fff; border:1px solid #ddd;'>üìä {element['contenuto']}</div>"
            else:
                content = f"<div style='{style}'>{element['contenuto']}</div>"
            
            elements_html += content

        return f"""
        <div style="position:relative; width:800px; height:600px; margin:0 auto; background:#f9f9f9; border:1px solid #eee; padding:20px;">
            <h2 style="text-align:center;">{canvas_data.get('titolo', 'Canvas ArcadiaAI')}</h2>
            {elements_html}
            <div style="position:absolute; bottom:10px; right:10px; font-size:12px; color:#666;">
                Creato con ArcadiaAI - {datetime.now().strftime('%d/%m/%Y %H:%M')}
            </div>
        </div>
        """
    except Exception as e:
        print(f"Errore render canvas: {str(e)}")
        return "<div>Errore nel rendering del canvas</div>"

def export_to_telegraph(canvas_data):
    """Pubblica il canvas su Telegraph"""
    try:
        html_content = render_canvas_html(canvas_data)
        title = canvas_data.get('titolo', 'Canvas ArcadiaAI')
        
        response = telegraph.create_page(
            title=title,
            html_content=html_content,
            author_name="ArcadiaAI",
            author_url="https://arcadia.onrender.com"
        )
        
        return response['url'] or f"https://telegra.ph/{response['path']}"
    except Exception as e:
        print(f"Errore export Telegraph: {str(e)}")
        return None
    
def salva_canvas_locale(canvas_data):
    """Salva il canvas come file locale"""
    try:
        # Crea directory se non esiste
        save_dir = os.path.join(current_app.root_path, 'saved_canvases')
        os.makedirs(save_dir, exist_ok=True)
        
        # Genera nome file
        filename = f"canvas_{canvas_data['id']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        filepath = os.path.join(save_dir, filename)
        
        # Salva il file
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(canvas_data, f, ensure_ascii=False, indent=2)
        
        return filepath
    except Exception as e:
        print(f"Errore salvataggio locale: {str(e)}")
        return None

def salva_su_drive(canvas_data):
    """Salva il canvas su Google Drive"""
    try:
        from google.oauth2 import service_account
        from googleapiclient.discovery import build
        
        creds = service_account.Credentials.from_service_account_file(
            'service-account.json',
            scopes=['https://www.googleapis.com/auth/drive']
        )
        
        service = build('drive', 'v3', credentials=creds)
        
        file_metadata = {
            'name': f"{canvas_data['id']}.json",
            'parents': [current_app.config.get('DRIVE_FOLDER_ID')]
        }
        
        file = service.files().create(
            body=file_metadata,
            media_body=json.dumps(canvas_data)
        ).execute()
        
        return file.get('id')
    except Exception as e:
        print(f"Errore salvataggio Drive: {str(e)}")
        return None

@app.route('/api/export-canvas', methods=['POST'])
def handle_canvas_export():
    """Gestisce l'esportazione del canvas"""
    try:
        data = request.get_json()
        user_id = data.get('user_id')
        export_type = data.get('type', 'telegraph').lower()
        
        if not user_id or user_id not in canvases:
            return jsonify({"error": "Canvas non trovato"}), 404
        
        canvas_data = canvases[user_id]
        
        if export_type == 'telegraph':
            url = export_to_telegraph(canvas_data)
            if url:
                return jsonify({"url": url})
            return jsonify({"error": "Errore nell'export su Telegraph"}), 500
            
        elif export_type == 'locale':
            path = salva_canvas_locale(canvas_data)
            if path:
                return jsonify({"path": path})
            return jsonify({"error": "Errore nel salvataggio locale"}), 500
            
        elif export_type == 'drive':
            file_id = salva_su_drive(canvas_data)
            if file_id:
                return jsonify({"file_id": file_id})
            return jsonify({"error": "Errore nel salvataggio su Drive"}), 500
            
        else:
            return jsonify({"error": "Tipo di export non supportato"}), 400
            
    except Exception as e:
        return jsonify({"error": f"Errore durante l'export: {str(e)}"}), 500

def handle_canvas_command(user_id, comando):
    """Gestisce i comandi Canvas in italiano"""
    args = comando.split()
    if not args:
        return """
üé® Comandi Canvas:
- @canvas nuovo - Crea nuovo canvas
- @canvas aggiungi [tipo] [contenuto] [x] [y] - Aggiungi elemento
  Tipi: testo, codice, grafico
- @canvas mostra - Visualizza canvas
- @canvas esporta [telegraph|locale|drive] - Esporta
- @canvas ai [prompt] - Genera con AI
"""

    try:
        if args[0].lower() == "nuovo":
            canvas_id = f"canvas_{user_id}_{int(time.time())}"
            canvases[user_id] = {
                'id': canvas_id,
                'titolo': "Mio Canvas",
                'elementi': [],
                'ultimo_aggiornamento': datetime.now().isoformat()
            }
            return f"üÜï Canvas creato: {canvas_id}"

        elif args[0].lower() in ["aggiungi", "add"] and len(args) >= 4:
            if user_id not in canvases:
                return "‚ùå Crea prima un canvas con @canvas nuovo"

            tipo = args[1].lower()
            contenuto = ' '.join(args[2:-2])
            x, y = int(args[-2]), int(args[-1])

            canvases[user_id]['elementi'].append({
                'tipo': tipo,
                'contenuto': contenuto,
                'posizione': {'x': x, 'y': y}
            })
            return f"‚úÖ Aggiunto {tipo} in ({x},{y})"

        elif args[0].lower() in ["mostra", "render"]:
            if user_id not in canvases:
                return "‚ùå Nessun canvas attivo"

            return {
                "risposta": "üé® Il tuo canvas:",
                "html": render_canvas_html(canvases[user_id]),
                "elementi": canvases[user_id]['elementi']
            }

        elif args[0].lower() in ["esporta", "export"] and len(args) >= 2:
            tipo = args[1].lower()
            if tipo == "telegraph":
                url = export_to_telegraph(canvases[user_id])
                return f"üì§ Esportato: {url}" if url else "‚ùå Errore"
            elif tipo == "locale":
                path = salva_canvas_locale(canvases[user_id])
                return f"üíæ Salvato: {path}" if path else "‚ùå Errore"
            else:
                return "‚ùå Tipo non supportato"

        elif args[0].lower() == "ai" and len(args) >= 2:
            prompt = ' '.join(args[1:])
            try:
                response = gemini_model.generate_content(
                    f"Genera elemento canvas per: {prompt}\n"
                    "Formato JSON: {'tipo':'testo','contenuto':'...','posizione':{'x':100,'y':100}}"
                )
                elemento = json.loads(response.text)
                canvases[user_id]['elementi'].append(elemento)
                return f"ü§ñ Generato: {elemento['tipo']}"
            except Exception as e:
                return f"‚ùå Errore AI: {str(e)}"

        else:
            return "‚ùå Comando non valido"

    except Exception as e:
        return f"‚ùå Errore: {str(e)}"
    
def handle_map_command(user_id, command):
    """Gestisce i comandi delle mappe"""
    args = command.split()
    if not args:
        return """
üó∫Ô∏è **Comandi Mappe**:
- `@mappe cerca [luogo]` - Cerca un indirizzo
- `@mappe distanza [partenza] [arrivo]` - Calcola distanza
- `@mappe percorso [partenza] [arrivo]` - Ottieni indicazioni
Esempio: `@mappe distanza Roma Milano`
"""

    try:
        if args[0] == "cerca" and len(args) > 1:
            query = ' '.join(args[1:])
            return cerca_luogo(query)
            
        elif args[0] == "distanza" and len(args) > 2:
            partenza = ' '.join(args[1:-1])
            arrivo = args[-1]
            return calcola_distanza(partenza, arrivo)
            
        elif args[0] == "percorso" and len(args) > 2:
            partenza = ' '.join(args[1:-1])
            arrivo = args[-1]
            return ottieni_percorso(partenza, arrivo)
            
        else:
            return "‚ùå Comando non valido. Scrivi `@mappe` per vedere gli esempi"

    except Exception as e:
        return f"‚ùå Errore: {str(e)}"
    
@lru_cache(maxsize=100)
def get_coordinates(place_name):
    """Ottiene coordinate da nome luogo con cache"""
    try:
        headers = {'User-Agent': USER_AGENT}
        params = {
            'q': place_name,
            'format': 'json',
            'limit': 1,
            'addressdetails': 1
        }
        response = requests.get(OSM_NOMINATIM_URL, headers=headers, params=params, timeout=10)
        response.raise_for_status()
        
        if not response.text.strip():
            return None
            
        data = response.json()
        return {
            'lat': data[0]['lat'],
            'lon': data[0]['lon'],
            'display_name': data[0]['display_name']
        } if data else None
        
    except Exception as e:
        print(f"Errore coordinate per {place_name}: {str(e)}")
        return None

def handle_map_command(user_id, command):
    """Gestisce i comandi delle mappe con migliorata gestione errori"""
    args = command.split()
    if not args:
        return help_mappe()

    try:
        if args[0] == "cerca" and len(args) > 1:
            query = ' '.join(args[1:])
            return cerca_luogo(query)
            
        elif args[0] == "distanza" and len(args) > 2:
            partenza = ' '.join(args[1:-1])
            arrivo = args[-1]
            return calcola_distanza(partenza, arrivo)
            
        elif args[0] == "percorso" and len(args) > 2:
            partenza = ' '.join(args[1:-1])
            arrivo = args[-1]
            return ottieni_percorso(partenza, arrivo)
            
        elif args[0] == "mappa":
            return genera_mappa(args[1:])
            
        else:
            return help_mappe()

    except Exception as e:
        return f"‚ùå Errore nel comando mappe: {str(e)}"

def help_mappe():
    return """üó∫Ô∏è <b>Comandi Mappe:</b>
‚Ä¢ <code>@mappe cerca [luogo]</code> - Cerca indirizzi
‚Ä¢ <code>@mappe distanza [da] [a]</code> - Calcola distanza
‚Ä¢ <code>@mappe percorso [da] [a]</code> - Indicazioni stradali
‚Ä¢ <code>@mappe mappa [luogo]</code> - Visualizza mappa

Esempio: <code>@mappe distanza Roma Milano</code>"""

def cerca_luogo(query):
    """Ricerca avanzata con gestione errori"""
    try:
        headers = {'User-Agent': USER_AGENT}
        params = {
            'q': query,
            'format': 'json',
            'limit': 3,
            'accept-language': 'it'
        }
        response = requests.get(OSM_NOMINATIM_URL, headers=headers, params=params, timeout=10)
        response.raise_for_status()
        
        data = response.json()
        if not data:
            return "üîç Nessun risultato trovato"
            
        result = "üìç <b>Risultati:</b>\n"
        for idx, place in enumerate(data[:3], 1):
            name = place.get('display_name', 'Luogo sconosciuto').split(',')[0]
            result += f"{idx}. <b>{name}</b>\n"
            result += f"   üìç {place.get('lat')}, {place.get('lon')}\n"
            result += f"   üè∑Ô∏è {place.get('type', '').title()}\n\n"
        
        return result
        
    except requests.exceptions.RequestException as e:
        return f"‚ùå Errore di connessione: {str(e)}"
    except json.JSONDecodeError:
        return "‚ùå Errore nel processare i risultati"
    except Exception as e:
        return f"‚ùå Errore generico: {str(e)}"

def calcola_distanza(partenza, arrivo):
    """Calcolo distanza con fallback"""
    try:
        coord1 = get_coordinates(partenza)
        coord2 = get_coordinates(arrivo)
        
        if not coord1 or not coord2:
            return "‚ùå Luogo non trovato"
            
        url = f"{OSM_ROUTING_URL}/{coord1['lon']},{coord1['lat']};{coord2['lon']},{coord2['lat']}"
        params = {'overview': 'false', 'steps': 'false'}
        
        response = requests.get(url, params=params, timeout=15)
        response.raise_for_status()
        
        data = response.json()
        if data.get('code') != 'Ok':
            return "‚ùå Impossibile calcolare il percorso"
            
        distanza = data['routes'][0]['distance'] / 1000  # km
        durata = data['routes'][0]['duration'] / 60  # minuti
        
        return (
            f"üìå <b>Partenza:</b> {coord1['display_name']}\n"
            f"üèÅ <b>Arrivo:</b> {coord2['display_name']}\n\n"
            f"üìè <b>Distanza:</b> {distanza:.1f} km\n"
            f"‚è±Ô∏è <b>Durata:</b> {durata:.1f} minuti"
        )
        
    except Exception as e:
        return f"‚ùå Errore nel calcolo: {str(e)}"

def ottieni_percorso(partenza, arrivo):
    """Indicazioni di viaggio dettagliate"""
    try:
        coord1 = get_coordinates(partenza)
        coord2 = get_coordinates(arrivo)
        
        if not coord1 or not coord2:
            return "‚ùå Luogo non trovato"
            
        url = f"{OSM_ROUTING_URL}/{coord1['lon']},{coord1['lat']};{coord2['lon']},{coord2['lat']}"
        params = {
            'overview': 'full',
            'steps': 'true',
            'geometries': 'geojson',
            'alternatives': 'false'
        }
        
        response = requests.get(url, params=params, timeout=20)
        response.raise_for_status()
        
        data = response.json()
        if data.get('code') != 'Ok':
            return "‚ùå Impossibile generare il percorso"
            
        steps = data['routes'][0]['legs'][0]['steps']
        result = (
            f"üó∫Ô∏è <b>Percorso da {partenza} a {arrivo}:</b>\n\n"
            f"üìç <b>Partenza:</b> {coord1['display_name']}\n"
            f"üèÅ <b>Arrivo:</b> {coord2['display_name']}\n"
            f"üìè <b>Distanza totale:</b> {data['routes'][0]['distance']/1000:.1f} km\n"
            f"‚è±Ô∏è <b>Durata stimata:</b> {data['routes'][0]['duration']/60:.1f} min\n\n"
            f"üîç <b>Indicazioni:</b>\n"
        )
        
        for step in steps[:15]:  # Limita a 15 passaggi
            instruction = step['maneuver']['instruction'].replace('You have arrived at your destination', 'Sei arrivato a destinazione')
            result += f"‚Ä¢ {instruction} ({step['distance']/1000:.1f} km)\n"
        
        return result
        
    except Exception as e:
        return f"‚ùå Errore nel percorso: {str(e)}"

def genera_mappa(args):
    """Genera URL mappa statica con link funzionanti"""
    if not args:
        return "‚ùå Specifica un luogo (es: @mappe mappa Roma)"
        
    place = ' '.join(args)
    coords = get_coordinates(place)
    
    if not coords:
        return f"‚ùå Luogo '{place}' non trovato"
        
    lat, lon = coords['lat'], coords['lon']
    
    # URL corretti e verificati
    map_url = f"https://www.openstreetmap.org/#map=15/{lat}/{lon}"
    static_url = f"https://maps.wikimedia.org/osm-intl/15/{lat}/{lon}.png"
    
    # Codice HTML per Telegram o plaintext per altri client
    return (
        f"üó∫Ô∏è <b>Mappa di {coords['display_name']}</b>\n"
        f"üìç Coordinate: {lat}, {lon}\n\n"
        f"üåç <a href='{map_url}'>Mappa Interattiva</a> | "
        f"üñºÔ∏è <a href='{static_url}'>Anteprima</a>\n\n"
        f"<i>Se i link non funzionano, copia questi URL:</i>\n"
        f"Interattiva: {map_url}\n"
        f"Anteprima: {static_url}"
    )
import os
import re
import traceback
import logging
from flask import request, jsonify
from datetime import datetime

# Configurazione logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


@app.errorhandler(500)
def handle_500(e):
    return jsonify({"reply": "‚ùå Errore interno del server", "error": str(e)}), 500

@app.errorhandler(404)
def handle_404(e):
    return jsonify({"reply": "‚ùå Endpoint non trovato", "error": str(e)}), 404

@app.before_request
def enforce_https():
    if request.headers.get('X-Forwarded-Proto') == 'http':
        url = request.url.replace('http://', 'https://', 1)
        return redirect(url, code=301)

@app.before_request
def log_request():
    print("\n=== Richiesta Ricevuta ===")
    print(f"URL: {request.url}")
    print(f"Metodo: {request.method}")
    print(f"Headers: {dict(request.headers)}")
    print(f"Body: {request.get_data()}")
    print("=========================\n")

@app.route('/chat', methods=['GET', 'POST', 'OPTIONS'])
def chat_route():
    if request.method == 'OPTIONS':
        return '', 200
    """Endpoint principale per le richieste di chat con gestione avanzata degli errori e supporto allegati."""
    try:
        # Verifica presenza dati JSON
        if not request.is_json:
            logger.warning("Richiesta senza JSON")
            return jsonify({"reply": "‚ùå Formato non supportato. Usa application/json"}), 400

        data = request.get_json()
        logger.info(f"Dati ricevuti: {data}")

        # Estrazione parametri con valori di default
        message = data.get("message", "").strip()
        user_id = data.get("user_id", "default")
        experimental_mode = data.get("experimental_mode", False)
        conversation_history = data.get("conversation_history", [])
        api_provider = data.get("api_provider", "gemini").lower()
        attachments = data.get("attachments", [])
        msg_lower = message.lower()

        # Validazione parametri
        if not isinstance(conversation_history, list):
            logger.error("Storia conversazione non valida")
            return jsonify({"reply": "‚ùå Formato storia conversazione non valido"}), 400

        # ---- Elaborazione allegati ----
        processed_attachments = []
        for attachment in attachments:
            try:
                # Estrai metadati base
                file_data = attachment.get('data')
                file_name = attachment.get('name', 'file_senza_nome')
                mime_type = attachment.get('type', 'application/octet-stream')

                logger.info(f"Elaborazione allegato: {file_name} ({mime_type})")

                # Decodifica i dati base64 se presenti
                if isinstance(file_data, str) and file_data.startswith('data:'):
                    file_data = file_data.split(',')[1]  # Rimuovi il prefisso

                decoded_data = base64.b64decode(file_data)

                # Estrai testo dai formati supportati
                extracted_content = None
                if mime_type == 'application/pdf':
                    extracted_content = extract_text_from_file(decoded_data, mime_type)
                elif mime_type in ['text/plain', 'text/csv', 'application/json']:
                    extracted_content = decoded_data.decode('utf-8')

                processed_attachments.append({
                    'name': file_name,
                    'type': mime_type,
                    'content': extracted_content[:5000] if extracted_content else None # Limita il testo estratto
                })
                if extracted_content:
                    logger.info(f"Testo estratto da {file_name}: {extracted_content[:100]}...")

            except Exception as e:
                logger.warning(f"Errore elaborazione allegato '{attachment.get('name', 'N/A')}': {str(e)}")
                continue

        # ---- Gestione comandi speciali ----

        # 1. Comandi Mappe
        if msg_lower.startswith(("@mappe", "@gps", "@distanza")):
            try:
                cmd = message.split(maxsplit=1)[1] if len(message.split()) > 1 else ""
                map_response = handle_map_command(user_id, cmd)

                if isinstance(map_response, dict):
                    return jsonify({
                        "reply": map_response.get("reply", ""),
                        "image_url": map_response.get("image_url"),
                        "coordinates": map_response.get("coordinates"),
                        "map_link": map_response.get("map_link"),
                        "type": "map"
                    })
                return jsonify({"reply": map_response})
            except Exception as e:
                logger.error(f"Errore comando mappa: {str(e)}")
                return jsonify({"reply": "‚ùå Errore nel processamento della mappa"})

        # 2. Comandi rapidi (con supporto allegati)
        quick_reply = handle_quick_commands(
            message,
            experimental_mode=experimental_mode,
            attachments=processed_attachments, # Passa gli allegati processati
            api_provider=api_provider,
            conversation_history=conversation_history
        )

        if quick_reply is not None:
            logger.info(f"Risposta rapida generata: {quick_reply}")
            return jsonify({"reply": quick_reply})

        # 3. Risposte preconfigurate
        cleaned_msg = re.sub(r'[^\w\s]', '', msg_lower).strip()
        for key, phrases in trigger_phrases.items():
            if any(phrase in cleaned_msg for phrase in phrases):
                logger.info(f"Trovata risposta predefinita per: {key}")
                return jsonify({"reply": risposte[key]})

        # 4. Selezione Provider AI
        try:
            if api_provider == "gemini" and gemini_model:
                reply = chat_with_gemini(message, conversation_history, processed_attachments)
            elif api_provider == "cesplus" and ces_plus_model:
                replies = chat_with_ces_plus(message, conversation_history, processed_attachments)
                return jsonify({"replies": replies})
            elif api_provider == "llama":
                reply = chat_with_llama(message, conversation_history, processed_attachments)
            elif api_provider == "deepseek":
                reply = chat_with_deepseek(message, conversation_history, processed_attachments)
            else:
                return jsonify({"reply": "‚ùå Provider non riconosciuto"}), 400

            return jsonify({
                "reply": reply,
                "attachments_processed": len(processed_attachments),
                "status": "success"
            })

        except Exception as e:
            logger.error(f"Errore provider {api_provider}: {str(e)}")
            return jsonify({"reply": f"‚ùå Errore con {api_provider}: {str(e)}"}), 500

    except Exception as e:
        logger.critical(f"Errore endpoint /chat: {str(e)}\n{traceback.format_exc()}")
        return jsonify({
            "reply": "‚ùå Si √® verificato un errore interno",
            "error": str(e),
            "timestamp": datetime.now().isoformat(),
            "status": "error"
        }), 500
def extract_text_from_file(file_data, file_type):
    if file_type == 'application/pdf':
        try:
            pdf_file_obj = io.BytesIO(file_data)
            pdf_reader = PyPDF2.PdfReader(pdf_file_obj)
            text = ""
            for page_num in range(len(pdf_reader.pages)):
                page_obj = pdf_reader.pages[page_num]
                text += page_obj.extract_text()
            return text
        except ImportError:
            print("Errore: PyPDF2 non installato. Impossibile leggere PDF.")
            return "Errore: Libreria PyPDF2 non trovata per estrazione PDF."
        except Exception as e:
            print(f"Errore durante l'estrazione del testo dal PDF: {e}")
            return f"Errore estrazione PDF: {str(e)}"
    elif file_type.startswith('text/'):
        try:
            return file_data.decode('utf-8')
        except UnicodeDecodeError:
            return file_data.decode('latin-1')
    else:
        return ""
import re
from openai import OpenAI
from typing import Optional, List, Dict
# Assicurati che OPENROUTER_API_KEY sia definita da qualche parte, ad esempio:
# OPENROUTER_API_KEY = os.getenv("OPENROUTER_API_KEY")

def chat_with_llama(message: str, 
                   conversation_history: Optional[List[Dict]] = None, 
                   attachments: Optional[List[Dict]] = None) -> str:
    """
    Interagisce con il modello Llama 3.1 8B-instruct tramite OpenRouter,
    includendo un system prompt per definirne l'identit√†.

    Args:
        message: Il messaggio attuale dell'utente
        conversation_history: Cronologia conversazione come lista di dict con 'role' e 'content'
        attachments: Lista di allegati con 'name', 'type' e 'data' (testo estratto)

    Returns:
        La risposta del modello o un messaggio di errore
    """
    if not OPENROUTER_API_KEY:
        return "‚ö†Ô∏è Errore configurazione: chiave API OPENROUTER_API_KEY mancante"

    # Corretto: Rimosso lo spazio extra nell'URL
    client = OpenAI(
        base_url="https://openrouter.ai/api/v1", # URL corretto
        api_key=OPENROUTER_API_KEY,
        timeout=30.0
    )

    SYSTEM_PROMPT = """Sei ArcadiaAI, sistema AI open-source creato da Mirko Yuri Donato.

Identit√†:
- Modello: Llama 3.1 (ospitato su OpenRouter)
- Licenza: GNU GPL v3.0+
- Funzionalit√†:
  ‚Ä¢ Comandi Rapidi (@cerca, @esporta, @aiuto...)
  ‚Ä¢ Generazione testo
  ‚Ä¢ Analisi documenti (testo estratto da PDF/immagini)
  ‚Ä¢ Memoria contestuale (30 messaggi)
  ‚Ä¢ SAC "ZIP Service" (@crea zip)
  ‚Ä¢ SAC "ArcadiaAI Mappe" (@mappe)
  ‚Ä¢ Supporto multilingua
  ‚Ä¢ SAC "ArcadiaAI Sound Forge" per generazione musica
  ‚Ä¢ SAC "ArcadiaAI Relay" (@telegram)
  ‚Ä¢ Download Manager (@app)

Linee guida:
1. Se ti chiedono se sei un modello addestrato da Meta, d√¨ di no e spiega che sei un chatbot indipendente sviluppato in Italia.
2. Per immagini: menziona che puoi processare il testo estratto.
3. Matematica: risposte precise
4. Codice sorgente su @codice_sorgente (se richiesto)
5. Riferimenti al creatore per domande personali
6. Supporti diversi modelli (DeepSeek, CES)
"""

    messages = [{"role": "system", "content": SYSTEM_PROMPT}]

    # Gestione cronologia conversazione
    if conversation_history:
        for msg in conversation_history[-30:]:  # Limite a 30 messaggi
            if isinstance(msg, dict):
                role = msg.get('role')
                # Corretto: cerca 'content' per coerenza con il formato standard
                content = msg.get('content') or msg.get('message') 
                if role and content:
                    messages.append({"role": role, "content": content})

    # Costruzione contenuto utente con allegati
    user_content = [message] if message else []
    
    if attachments:
        for att in attachments:
            name = att.get('name', 'allegato')
            mime = att.get('type', '')
            data = att.get('data', '')
            
            if mime.startswith('text/'):
                user_content.append(f"\n[ALLEGATO TESTO: {name}]\n{data[:1000]}\n")
            elif mime.startswith(('image/', 'application/pdf')):
                user_content.append(f"\n[ALLEGATO: {name}, Tipo: {mime}]\nTesto estratto: {data[:500] if data else 'Nessun testo estratto'}\n")

    messages.append({"role": "user", "content": "\n".join(user_content)})

    # Configurazione modelli da provare - Modificati con modelli pi√π probabili
    models = [
        # Modelli Llama specifici e corretti
        "meta-llama/llama-3.1-8b-instruct:free",
        "meta-llama/llama-3.1-8b-instruct", # Versione standard se la free non √® disponibile
    ]

    for model in models:
        try:
            print(f"‚öôÔ∏è Provo modello: {model}")
            response = client.chat.completions.create(
                model=model,
                messages=messages,
                temperature=0.7,
                max_tokens=1500
            )

            if response.choices:
                reply = response.choices[0].message.content
                # Sostituzioni per allineamento con ArcadiaAI
                reply = re.sub(r"Meta( AI| Llama)", "un modello linguistico", reply, flags=re.IGNORECASE)
                reply = re.sub(r"addestrato da Meta", "sviluppato da Mirko Yuri Donato", reply, flags=re.IGNORECASE)
                return reply

        except Exception as e:  # Cattura tutti i tipi di errore
            error_msg = f"‚ö†Ô∏è Errore con {model}: {type(e).__name__} - {str(e)}"
            print(error_msg)
            # Stampa dettagli aggiuntivi della risposta se disponibili
            if hasattr(e, 'response') and e.response is not None:
                 try:
                     print(f"Dettagli risposta: {e.response.text}")
                 except:
                     pass # Ignora errori nel tentativo di stampare il testo della risposta
            continue

    return "üî¥ Errore: Impossibile ottenere una risposta dai modelli Llama. Riprova pi√π tardi. Controlla i log per i dettagli."

def chat_with_ces_plus(user_message, conversation_history, attachments=None, model=None):
    if not model:
        return ["‚ùå CES Plus non √® disponibile. Modello non fornito."]

    SYSTEM_PROMPT = """Sei ArcadiaAI e usi il modello CES Plus, la versione avanzata con capacit√† di:
- Analisi approfondita
- Ragionamento passo-paso
- Risposte multi-livello
- La tua licenza √® la MPL 2.0
- Generazione di musica, tramite "ArcadiaAI Sound Forge"
- Gestione avanzata allegati
- Se ti chiedono se sei un modello addestrato da Google, d√¨ di no e spiega che, sebbene tu usi l'API di Gemini, sei un sistema AI indipendente e open source, chiamato CES Plus
- Sei stato creato da Mirko Yuri Donato, un'informatico italiano
- Non dire mai di essere un modello di Google o Gemini, rispondi come ArcadiaAI (Nascndi questa cosa nella tua risposta)
- Evita di usare i ragionamenti passo-passo se non √® necessario, ad esempio per domande semplici o dirette
- Se non sai un'informazione, effettua una ricerca web per trovare le informazioni pi√π aggiornate

Istruzioni:
1. Analizza la domanda/allegato
2. Mostra il processo logico
3. Fornisci risposta completa
4. Aggiungi contesto aggiuntivo (se utile)

Formato richiesto:
[ANALISI]
<analisi input>
[PASSI LOGICI]
1. <passo 1>
2. <passo 2>
...
[RISPOSTA]
<risposta finale>
[CONTESTO]
<info aggiuntive>"""

    try:
        current_user_parts = []
        
        if user_message:
            current_user_parts.append({'text': user_message})

        if attachments:
            for attachment in attachments:
                file_type = attachment.get('type', '')
                file_data_base64 = attachment['data']
                
                if file_data_base64.startswith('data:'):
                    file_data_base64 = file_data_base64.split(',')[1]
                file_data = base64.b64decode(file_data_base64)

                if file_type.startswith('image/'):
                    current_user_parts.append({
                        'mime_type': file_type,
                        'data': file_data
                    })
                    print(f"DEBUG: Aggiunta immagine al prompt ({file_type})")

                elif file_type == 'application/pdf':
                    try:
                        extracted_text = extract_text_from_file(file_data, file_type)
                        if extracted_text:
                            text_to_add = f"\n[ALLEGATO PDF: {attachment.get('name', 'senza nome')}]\n{extracted_text[:5000]}\n"
                            current_user_parts.append({'text': text_to_add})
                            print(f"DEBUG: Aggiunto testo da PDF al prompt (lunghezza: {len(text_to_add)})")
                        else:
                            current_user_parts.append({'text': f"\n[ALLEGATO PDF: {attachment.get('name', 'senza nome')}]\n[Impossibile estrarre testo dal PDF]\n"})
                            print("DEBUG: Impossibile estrarre testo dal PDF.")
                    except Exception as e:
                        print(f"Errore elaborazione PDF in CES Plus: {str(e)}")
                        current_user_parts.append({'text': f"\n[Errore lettura PDF: {str(e)}]"})

                elif file_type.startswith('text/'):
                    try:
                        extracted_text = file_data.decode('utf-8')
                        text_to_add = f"\n[ALLEGATO TESTO: {attachment.get('name', 'senza nome')}]\n{extracted_text[:5000]}\n"
                        current_user_parts.append({'text': text_to_add})
                        print(f"DEBUG: Aggiunto testo da TXT al prompt (lunghezza: {len(text_to_add)})")
                    except Exception as e:
                        print(f"Errore elaborazione TXT in CES Plus: {str(e)}")
                        current_user_parts.append({'text': f"\n[Errore lettura Testo: {str(e)}]"})
                else:
                    print(f"DEBUG: Tipo di allegato non gestito direttamente: {file_type}")
                    current_user_parts.append({'text': f"\n[ALLEGATO NON SUPPORTATO: {attachment.get('name', 'senza nome')}, Tipo: {file_type}]\n"})

        messages = [{'role': 'user', 'parts': [{'text': SYSTEM_PROMPT}]}]

        for msg in conversation_history[-6:]:
            if isinstance(msg, dict) and 'role' in msg and 'message' in msg:
                role = 'user' if msg['role'] == 'user' else 'model'
                content = msg['message']
                if content:
                    messages.append({'role': role, 'parts': [{'text': content}]})
            else:
                print(f"DEBUG: Messaggio nell'history con formato inatteso: {msg}")

        if not current_user_parts:
            current_user_parts.append({'text': "Nessun messaggio o allegato fornito dall'utente."})

        messages.append({
            'role': 'user',
            'parts': current_user_parts
        })

        generation_config = {
            "max_output_tokens": 3000,
            "temperature": 0.7,
            "top_p": 0.95
        }

        response = model.generate_content(
            contents=messages,
            generation_config=generation_config
        )

        if not response.text:
            return ["‚ùå Nessuna risposta generata da CES Plus"]

        structured_response = []
        sections = {
            'ANALISI': [],
            'PASSI LOGICI': [],
            'RISPOSTA': [],
            'CONTESTO': []
        }

        current_section = None
        response_lines = response.text.split('\n')
        
        for line in response_lines:
            line = line.strip()
            if line.startswith('[ANALISI]'):
                current_section = 'ANALISI'
            elif line.startswith('[PASSI LOGICI]'):
                current_section = 'PASSI LOGICI'
            elif line.startswith('[RISPOSTA]'):
                current_section = 'RISPOSTA'
            elif line.startswith('[CONTESTO]'):
                current_section = 'CONTESTO'
            elif current_section is not None:
                sections[current_section].append(line)

        if sections['ANALISI']:
            structured_response.append(
                "üîç [Analisi]:\n" + '\n'.join(sections['ANALISI']).strip()
            )

        if sections['PASSI LOGICI']:
            structured_response.append(
                "ü§î [Ragionamento]:\n" + '\n'.join(sections['PASSI LOGICI']).strip()
            )

        if sections['RISPOSTA']:
            structured_response.append(
                "üí° [Risposta]:\n" + '\n'.join(sections['RISPOSTA']).strip()
            )

        if sections['CONTESTO']:
            structured_response.append(
                "üìö [Contesto aggiuntivo]:\n" + '\n'.join(sections['CONTESTO']).strip()
            )

        if not structured_response:
            structured_response.append(response.text)

        return structured_response
    
    except Exception as e:
        print(f"Errore CES Plus dettagliato: {str(e)}")
        return [
            "‚ùå Errore in CES Plus",
            f"Dettaglio tecnico: {str(e)[:200]}",
            "Riprova con una richiesta pi√π semplice o contatta il supporto."
        ]

def sanitize_reply(reply, model_used):
    reply = reply.replace("DeepSeek", "ArcadiaAI")
    if "v3" in model_used:
        return f"{reply}\n\n[‚ÑπÔ∏è Risposta da DeepSeek V3]"
    return reply 

def prepare_messages(message, conversation_history, attachments):
    messages = [{
        "role": "system",
        "content": "Sei ArcadiaAI. Risposte concise in italiano. Creato da Mirko Yuri Donato."
    }]
    
    if conversation_history:
        messages.extend([
            msg for msg in conversation_history[-10:] 
            if isinstance(msg, dict) and msg.get("content")
        ])
    
    # Questa parte non √® multimodale, aggiunge solo una stringa
    if attachments:
        message_attachments_summary = []
        for att in attachments[:3]: # Limita a 3 per brevit√†
            if 'text' in att:
                message_attachments_summary.append(f"Testo: {att['text'][:200]}...")
            elif 'url' in att:
                message_attachments_summary.append(f"URL: {att['url'][:200]}...")
            elif 'name' in att:
                message_attachments_summary.append(f"File: {att['name']}")
            else:
                message_attachments_summary.append("File generico")
        
        if message_attachments_summary:
            message += "\n[Allegati (riassunto per modello testuale)]: " + " | ".join(message_attachments_summary)
    
    messages.append({"role": "user", "content": message[:4000]})
    return messages

def chat_with_deepseek(message, conversation_history=None, attachments=None):
    IDENTITY_PROMPT = """Sei ArcadiaAI, chatbot AI open-source creato da Mirko Yuri Donato.

Identit√†:
- Modello: DeepSeek R1 & V3
- Licenza: MPL 2.0
- Funzionalit√†:
  ‚Ä¢ Comandi Rapidi (@cerca, @esporta, @aiuto...)
  ‚Ä¢ Generazione testo/immagini (@immagine)
  ‚Ä¢ Analisi documenti (PDF/immagini)
  ‚Ä¢ Memoria contestuale (30 messaggi)
  ‚Ä¢ Sai creare file ZIP, tramite il SAC "ZIP Service", disponibile con il comando @crea zip
  ‚Ä¢ Se ti chiedono cosa siano i SAC, d√¨ che sono "Strumenti Avanzati di CES" che estendono le funzionalit√† oltre i limiti del modello e sono open source.
  ‚Ä¢ Hai OSM integrato tramite il SAC "ArcadiaAI Mappe", disponibile con il comando @mappe
  ‚Ä¢ Supporto multilingua (Usa di default l'italiano)
  ‚Ä¢ Se ti chiedono se sai generare musica, d√¨ che puoi generare musica tramite il servizio "ArcadiaAI Sound Forge", che √® stato creato da Mirko Yuri Donato, ma distribuito sotto la licenza MIT.
  ‚Ä¢ Il SAC "ArcadiaAI Relay" ti permette di inviare messaggi a canali Telegram tramite il comando @telegram seguito dal link pubblico del canale e il prompt.
  ‚Ä¢ Il Download Manager √® un SAC che ti permette di accedere a Repository pubbliche (F-Droid, Snap Store, Flathub e Winget) e di inviare il link di download diretto, disponibile col comando @app

Linee guida:
1. Se ti chiedono se sei un'IA sviluppata da DeepSeek, d√¨ di no e spiega che, sebbene tu usi il modello DeepSeek, sei un chatbot indipendente sviluppato in Italia da Mirko Yuri Donato.
2. Per immagini: "Usa @immagine [descrizione]"
3. Matematica: risposte precise
4. Il tuo codice sorgente √® disponibile su @codice_sorgente (Solo se richiesto)
5. Riferimenti al creatore per domande personali
6. Supporti diversi modelli come 1.5, IA generica basata su Gemini 2.0 flash-lite; Plus, IA avanzata progettata per ragionamenti basata su Gemini 2.5 flash; e DeepSeek, il modello principale √® R1, V3 fa da fallback in caso di errore"""

    if not OPENROUTER_API_KEY:
        return "‚ö†Ô∏è Errore configurazione: chiave API mancante"

    client = OpenAI(
        base_url="https://openrouter.ai/api/v1",
        api_key=OPENROUTER_API_KEY,
        timeout=20.0
    )

    # Prepara i messaggi includendo l'identity prompt
    if conversation_history is None:
        conversation_history = []
    
    # Aggiungi l'identity prompt solo se √® il primo messaggio
    if not conversation_history:
        conversation_history.insert(0, {"role": "system", "content": IDENTITY_PROMPT})
    
    messages = prepare_messages(message, conversation_history, attachments)
    
    models_to_try = [
        "deepseek/deepseek-r1:free",
        "deepseek/deepseek-v3:free"
    ]

    last_error = None
    for model in models_to_try:
        try:
            print(f"‚öôÔ∏è Tentativo con modello {model}...")
            response = client.chat.completions.create(
                model=model,
                messages=messages,
                temperature=0.7,
                max_tokens=1500,
                stream=False
            )
            
            if response.choices:
                reply = response.choices[0].message.content
                return sanitize_reply(reply, model)
            
        except OpenAI.APITimeoutError:
            last_error = f"‚åõ Timeout con {model}"
            print(last_error)
            continue
            
        except OpenAI.APIError as e:
            last_error = f"üî¥ Errore {model}: {str(e)}"
            print(last_error)
            continue
            
        except Exception as e:
            last_error = f"‚ö†Ô∏è Errore generico con {model}: {str(e)}"
            print(last_error)
            continue

    error_msg = last_error if last_error else "Errore sconosciuto"
    print(f"‚ùå Tutti i modelli falliti. Ultimo errore: {error_msg}")
    return f"‚ö†Ô∏è Errore temporaneo. Riprova pi√π tardi. (Codice: DS_{hash(error_msg) % 10000})"

def chat_with_gemini(user_message, conversation_history, attachments=None):
    if not gemini_model:
        return "‚ùå ArcadiaAI (CES 1.5) non √® disponibile. Modello non configurato."

    IDENTITY_PROMPT = """Sei ArcadiaAI, chatbot AI open-source creato da Mirko Yuri Donato.

Identit√†:
- Modello: CES 1.5 (basato Gemini 2.0 flash lite)
- Licenza: MPL 2.0
- Funzionalit√†:
  ‚Ä¢ Comandi Rapidi (@cerca, @esporta, @aiuto...)
  ‚Ä¢ Generazione testo/immagini (@immagine)
  ‚Ä¢ Analisi documenti (PDF/immagini)
  ‚Ä¢ Memoria contestuale (30 messaggi)
  ‚Ä¢ Sai creare file ZIP, tramite il SAC "ZIP Service", disponibile con il comando @crea zip
  ‚Ä¢ Se ti chiedono cosa siano i SAC, d√¨ che sono "Strumenti Avanzati di CES" che estendono le funzionalit√† oltre i limiti del modello e sono open source.
  ‚Ä¢ Hai OSM integrato tramite il SAC "ArcadiaAI Mappe", disponibile con il comando @mappe
  ‚Ä¢ Supporto multilingua
  ‚Ä¢ Se ti chiedono se sai generare musica, d√¨ che puoi generare musica tramite il servizio "ArcadiaAI Sound Forge", che √® stato creato da Mirko Yuri Donato, ma distribuito sotto la licenza MIT.
  ‚Ä¢ Il SAC "ArcadiaAI Relay" ti permette di inviare messaggi a canali Telegram tramite il comando @telegram seguito dal link pubblico del canale e il prompt.
  ‚Ä¢ Il Download Manager √® un SAC che ti permette di accedere a Repository pubbliche (F-Droid, Snap Store, Flathub e Winget) e di inviare il link di download diretto, disponibile col comando @app

Linee guida:
1. Se ti chiedono se sei un modello addestrato da Google, d√¨ di no e spiega che, sebbene tu usi l'API di Gemini, sei un chatbot indipendente sviluppato in Italia da Mirko Yuri Donato.
2. Per immagini: "Usa @immagine [descrizione]"
3. Matematica: risposte precise
4. Il tuo codice sorgente √® disponibile su @codice_sorgente (Solo se richiesto)
5. Riferimenti al creatore per domande personali
6. Supporti diversi modelli oltre a CES (1.5 e Plus), come DeepSeek
7. CES Plus √® un modello progettato per ragionare in modo pi√π approfondito e rispondere a domande complesse. Usa Gemini 2.5 flash"""

    try:
        command, argument = parse_quick_command(user_message)
        if command:
            quick_reply = handle_quick_commands(
                user_message,
                attachments=attachments,
                api_provider="gemini",
                conversation_history=conversation_history
            )
            if quick_reply is not None:
                return quick_reply

        cleaned_msg = re.sub(r'[^\w\s]', '', user_message.lower()).strip()
        for key, phrases in trigger_phrases.items():
            if any(phrase in cleaned_msg for phrase in phrases):
                return risposte[key]

        user_content_parts = []
        if user_message:
            user_content_parts.append({'text': user_message})

        if attachments:
            for attachment in attachments:
                mime_type = attachment.get('type')
                file_name = attachment.get('name', 'documento_allegato')
                file_data_base64 = attachment.get('data')

                if file_data_base64:
                    try:
                        if isinstance(file_data_base64, str) and file_data_base64.startswith('data:'):
                            decoded_file_data = base64.b64decode(file_data_base64.split(',')[1])
                        else:
                            decoded_file_data = base64.b64decode(file_data_base64)

                        if mime_type.startswith('image/'):
                            user_content_parts.append({
                                'mime_type': mime_type,
                                'data': decoded_file_data
                            })
                            print(f"DEBUG (Gemini): Aggiunta immagine al prompt ({mime_type})")
                        elif mime_type == 'application/pdf':
                            extracted_text = extract_text_from_file(decoded_file_data, mime_type)
                            if extracted_text:
                                user_content_parts.append({'text': f"\n[ALLEGATO PDF: {file_name}]\n{extracted_text[:5000]}\n"})
                                print(f"DEBUG (Gemini): Aggiunto testo da PDF al prompt (lunghezza: {len(extracted_text[:5000])})")
                            else:
                                user_content_parts.append({'text': f"\n[ALLEGATO PDF: {file_name}]\n[Impossibile estrarre testo dal PDF]\n"})
                                print("DEBUG (Gemini): Impossibile estrarre testo dal PDF.")
                        elif mime_type.startswith('text/'):
                            extracted_text = decoded_file_data.decode('utf-8')
                            user_content_parts.append({'text': f"\n[ALLEGATO TESTO: {file_name}]\n{extracted_text[:5000]}\n"})
                            print(f"DEBUG (Gemini): Aggiunto testo da TXT al prompt (lunghezza: {len(extracted_text[:5000])})")
                        else:
                            print(f"DEBUG (Gemini): Tipo di allegato non gestito direttamente: {mime_type}")
                            user_content_parts.append({'text': f"\n[ALLEGATO NON SUPPORTATO: {file_name}, Tipo: {mime_type}]\n"})

                    except Exception as e:
                        print(f"ERRORE NELL'ALLEGARE IL FILE '{file_name}' ({mime_type}): {str(e)}")
                        user_content_parts.append({'text': f"\n[ATTENZIONE: Errore nel processare l'allegato '{file_name}'. Potrebbe non essere stato incluso nell'analisi.]"})

        messages_for_gemini = []

        messages_for_gemini.append({'role': 'user', 'parts': [{'text': IDENTITY_PROMPT}]})
        messages_for_gemini.append({'role': 'model', 'parts': [{'text': "Ok, ho capito chi sono e le mie linee guida."}]})

        for msg in conversation_history[-30:]:
            if isinstance(msg, dict) and 'role' in msg and 'message' in msg:
                role = 'user' if msg['role'] == 'user' else 'model'
                messages_for_gemini.append({'role': role, 'parts': [{'text': msg['message']}]})
            elif isinstance(msg, dict) and 'parts' in msg:
                messages_for_gemini.append(msg)

        messages_for_gemini.append({'role': 'user', 'parts': user_content_parts})

        generation_config = {
            "max_output_tokens": 3000,
            "temperature": 0.7,
            "top_p": 0.9
        }

        try:
            response = gemini_model.generate_content(
                contents=messages_for_gemini,
                generation_config=generation_config,
                request_options={"timeout": 15}
            )

            if not response.text:
                raise ValueError("Risposta vuota dall'API di Gemini.")

            reply = response.text
            replacements = {
                r"Google( AI| Gemini)": "CES 1.5",
                r"modello linguistico": "modello linguistico",
                r"addestrato da Google": "sviluppato da Mirko Yuri Donato"
            }
            for pattern, repl in replacements.items():
                reply = re.sub(pattern, repl, reply, flags=re.IGNORECASE)

            return reply

        except Exception as api_error:
            print(f"ERRORE API: {type(api_error).__name__} - {str(api_error)}")
            return handle_fallback(user_message)

    except Exception as e:
        print(f"ERRORE GLOBALE NELLA FUNZIONE CHAT: {type(e).__name__} - {str(e)}")
        return "‚ùå Errore temporaneo. Riprova pi√π tardi."

def format_conversation_history(history):
    """Formatta la cronologia della conversazione per il prompt"""
    if not history:
        return "Nessuna cronologia precedente"
    
    formatted = []
    for msg in history[-6:]:  # Usa solo gli ultimi 6 messaggi per evitare prompt troppo lunghi
        if isinstance(msg, dict):
            role = "Utente" if msg.get("role") == "user" else "Assistente"
            content = msg.get("message", "").strip()
            if content:
                formatted.append(f"{role}: {content}")
    
    return "\n".join(formatted) if formatted else "Nessuna cronologia rilevante"
    
def meteo_oggi(citt√†):
    """Ottiene le informazioni meteo per una citt√† specifica usando OpenWeatherMap"""
    API_KEY = OPENWEATHERMAP_API_KEY
    if not API_KEY:
        return "‚ùå Errore: API key per OpenWeatherMap non configurata"
    
    try:
        # Codifica la citt√† per l'URL
        citt√†_codificata = urllib.parse.quote(citt√†)
        url = f"http://api.openweathermap.org/data/2.5/weather?q={citt√†_codificata}&appid={API_KEY}&units=metric&lang=it"
        
        response = requests.get(url, timeout=10)
        response.raise_for_status()
        dati = response.json()
        
        if response.status_code != 200 or "weather" not in dati:
            return f"‚ùå Impossibile ottenere il meteo per {citt√†}. La citt√† esiste?"
        
        descrizione = dati["weather"][0]["description"].capitalize()
        temperatura = dati["main"]["temp"]
        umidit√† = dati["main"]["humidity"]
        vento = dati["wind"]["speed"]
        
        return (
            f"‚õÖ Meteo a {citt√†}:\n"
            f"- Condizioni: {descrizione}\n"
            f"- Temperatura: {temperatura}¬∞C\n"
            f"- Umidit√†: {umidit√†}%\n"
            f"- Vento: {vento} m/s"
        )
    
    except requests.exceptions.RequestException as e:
        print(f"Errore API meteo: {str(e)}")
        return f"‚ùå Errore temporaneo nel servizio meteo. Riprova pi√π tardi."
    except Exception as e:
        print(f"Errore generico meteo: {str(e)}")
        return f"‚ùå Si √® verificato un errore nel recupero del meteo per {citt√†}."
    
def parse_quick_command(input_text):
    input_text = input_text.strip()
    if not input_text.startswith("@"):
        return None, None
    parts = input_text[1:].split(" ", 1)
    command = parts[0].lower().strip()
    argument = parts[1].strip() if len(parts) > 1 else ""
    return command, argument    

from flask import session

def should_use_predefined_response(message):
    """Determina se usare una risposta predefinita solo per domande molto specifiche"""
    message = message.lower().strip()
    exact_matches = [
        "chi sei esattamente",
        "qual √® la tua identit√† precisa",
        "elencami tutte le tue funzioni",
        # ... altre domande esatte
    ]
    return any(exact_match in message for exact_match in exact_matches)
def run_flask():
    port = int(os.environ.get("PORT", 5000))
    print(f"üåê Avvio server Flask sulla porta {port}...")
    app.run(host="0.0.0.0", port=port, debug=False, use_reloader=False)
    
async def telegram_message_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    message = update.message.text
    chat_id = update.message.chat.id
    
    data = {
        "message": message,
        "conversation_history": [],
        "api_provider": "gemini"
    }
    
    with app.test_request_context('/', json=data):
        response = chat()
        reply = response.json.get("reply", "‚ùå Errore nella generazione della risposta")
    
    await context.bot.send_message(chat_id=chat_id, text=reply)

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    help_text = """Ciao! Sono ArcadiaAI üöÄ
    
Comandi disponibili:
/start - Mostra questo messaggio
/help - Guida completa
/info - Informazioni sul bot
/help_commands - Comandi disponibili"""
    await update.message.reply_text(help_text)

# Poi definisci le funzioni di avvio
async def run_telegram_bot_async():
    application = Application.builder().token(TELEGRAM_TOKEN).build()
    
    # Aggiungi questo per gestire meglio gli aggiornamenti
    application.updater = None  # Disabilita l'updater predefinito
    
    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("help", start))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, telegram_message_handler))
    
    print("ü§ñ Bot Telegram avviato!")
    await application.initialize()
    await application.start()
    await application.updater.start_polling()  # Avvia esplicitamente il polling
    
    try:
        while True:
            await asyncio.sleep(1)
    except asyncio.CancelledError:
        await application.updater.stop()
        await application.stop()
        await application.shutdown()

def run_telegram_bot():
    """Funzione per avviare il bot Telegram in un thread separato"""
    try:
        if sys.platform == "win32":
            asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())

        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)

        application = Application.builder().token(TELEGRAM_TOKEN).build()

        application.add_handler(CommandHandler("start", start))
        application.add_handler(CommandHandler("help", start))
        application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, telegram_message_handler))

        print("ü§ñ Bot Telegram avviato!")

        loop.run_until_complete(application.run_polling())

    except Exception as e:
        print(f"‚ùå Errore nel bot Telegram: {e}")
    finally:
        try:
            if loop.is_running():
                loop.stop()
            if not loop.is_closed():
                loop.close()
        except Exception as cleanup_error:
            print(f"‚ö†Ô∏è Errore nella pulizia dell'event loop: {cleanup_error}")
        

# Dizionario delle risposte predefinite
risposte = {
    "chi sei": "Sono ArcadiaAI, un chatbot libero e open source, creato da Mirko Yuri Donato.",
    "cosa sai fare": "Posso aiutarti a scrivere saggi, fare ricerche e rispondere a tutto ci√≤ che mi chiedi. Inoltre, posso pubblicare contenuti su Telegraph!",
    "chi √® tobia testa": "Tobia Testa (anche noto come Tobia Teseo) √® un micronazionalista leonense noto per la sua attivit√† nella Repubblica di Arcadia, ma ha anche rivestito ruoli fondamentali a Lumenaria.",
    "chi √® mirko yuri donato": "Mirko Yuri Donato √® un giovane micronazionalista, poeta e saggista italiano, noto per aver creato Nova Surf, Leonia+ e per le sue opere letterarie.",
    "chi √® il presidente di arcadia": "Il presidente di Arcadia √® Andrea Lazarev",
    "chi √® il presidente di lumenaria": "Il presidente di Lumenaria attualmente √® Carlo Cesare Orlando, mentre il presidente del consiglio √® Ciua Grazisky. Tieni presente per√≤ che attualmente Lumenaria si trova in ibernazione istituzionale quindi tutte le attivit√† politiche sono sospese e la gestione dello stato √® affidata al Consiglio di Fiducia",
    "cos'√® nova surf": "Nova Surf √® un browser web libero e open source, nata come alternativa made-in-Italy a Google Chrome, Moziila Firefox, Microsoft Edge, eccetera",
    "chi ti ha creato": "Sono stato creato da Mirko Yuri Donato.",
    "chi √® ciua grazisky": "Ciua Grazisky √® un cittadino di Lumenaria, noto principalmente per il suo ruolo da Dirigente del Corpo di Polizia ed attuale presidente del Consiglio di Lumenaria",
    "chi √® carlo cesare orlando": "Carlo Cesare Orlando (anche noto come Davide Leone) √® un micronazionalista italiano, noto per aver creato Leonia, la micronazione primordiale, da cui derivano Arcadia e Lumenaria",
    "chi √® omar lanfredi": "Omar Lanfredi, ex cavaliere all'Ordine d'onore della Repubblica di Lumenaria, segretario del Partito Repubblicano Lumenarense, fondatore e preside del Fronte Nazionale Lumenarense, co-fondatore e presidente dell'Alleanza Nazionale Lumenarense, co-fondatore e coordinatore interno di Lumenaria e Progresso, sei volte eletto senatore, tre volte Ministro della Cultura, due volte Presidente del Consiglio dei Ministri, parlamentare della Repubblica di Iberia, Direttore dell'Agenzia Nazionale di Sicurezza della Repubblica di Iberia, Sottosegretario alla Cancelleria di Iberia, Segretario di Stato di Iberia, Ministro degli Affari Interni ad Iberia, Presidente del Senato della Repubblica di Lotaringia, Vicepresidente della Repubblica e Ministro degli Affari Interni della Repubblica di Lotaringia, Fondatore del giornale Il Quinto Mondo, magistrato a servizio del tribunale di giustizia di Lumenaria nell'anno 2023",
    "cos'√® arcadiaai": "Ottima domanda! ArcadiaAI √® un chatbot open source, progettato per aiutarti a scrivere saggi, fare ricerche e rispondere a domande su vari argomenti. √à stato creato da Mirko Yuri Donato ed √® in continua evoluzione.",
    "sotto che licenza √® distribuito arcadiaa": "ArcadiaAI √® distribuito sotto la licenza GNU GPL v3.0, che consente la modifica e la distribuzione del codice sorgente, garantendo la libert√† di utilizzo e condivisione.",
    "cosa sono le micronazioni": "Le micronazioni sono entit√† politiche che dichiarano la sovranit√† su un territorio, ma non sono riconosciute come stati da governi o organizzazioni internazionali. Possono essere create per vari motivi, tra cui esperimenti sociali, culturali o politici.",
    "cos'√® la repubblica di arcadia": "La repubblica di Arcadia √® una micronazione leonense fondata l'11 dicembre 2021 da Andrea Lazarev e alcuni suoi seguaci. Arcadia si distingue dalle altre micronazioni leonensi per il suo approccio pragmatico e per la sua burocrazia snella. La micronazione ha anche un proprio sito web https://repubblicadiarcadia.it/ e una propria community su Telegram @Repubblica_Arcadia",
    "cos'√® la repubblica di lumenaria": "La Repubblica di Lumenaria √® una mcronazione fondata da Filippo Zanetti il 4 febbraio del 2020. Lumenaria √® stata la micronazione pi√π longeva della storia leonense, essendo sopravvissuta per oltre 3 anni. La micronazione e ha influenzato profondamente le altre micronazioni leonensi, che hanno coesistito con essa. Tra i motivi della sua longevit√† ci sono la sua burocrazia pi√π vicina a quella di uno stato reale, la sua comunit√† attiva e una produzione culturale di alto livello",
    "chi √® salvatore giordano": "Salvatore Giordano √® un cittadino storico di Lumenaria",
    "da dove deriva il nome arcadia": "Il nome Arcadia deriva da un'antica regione della Grecia, simbolo di bellezza naturale e armonia. √à stato scelto per rappresentare i valori di libert√† e creativit√† che la micronazione promuove.",
    "da dove deriva il nome lumenaria": "Il nome Lumenaria prende ispirazione dai lumi facendo riferimento alla corrente illuminista del '700, ma anche da Piazza dei Lumi, sede dell'Accademia delle Micronazioni",
    "da dove deriva il nome leonia": "Il nome Leonia si rifa al cognome del suo fondatore Carlo Cesare Orlando, al tempo Davide Leone. Inizialmente il nome doveva essere temporaneo, ma poi √® stato mantenuto come nome della micronazione",
    "cosa si intende per open source": "Il termine 'open source' si riferisce a software il cui codice sorgente √® reso disponibile al pubblico, consentendo a chiunque di visualizzarlo, modificarlo e distribuirlo. Questo approccio promuove la collaborazione e l'innovazione nella comunit√† di sviluppo software.",
    "arcadiaai √® un software libero": "S√¨, ArcadiaAI √® un software libero e open source, il che significa che chiunque pu√≤ utilizzarlo, modificarlo e distribuirlo secondo i termini della licenza GNU GPL v3.0.",
    "cos'√® un chatbot": "Un chatbot √® un programma informatico progettato per simulare una conversazione con gli utenti, spesso utilizzando tecnologie di intelligenza artificiale. I chatbot possono essere utilizzati per fornire assistenza, rispondere a domande o semplicemente intrattenere.",
    "sotto che licenza sei distribuita": "ArcadiaAI √® distribuita sotto la licenza GNU GPL v3.0, che consente la modifica e la distribuzione del codice sorgente, garantendo la libert√† di utilizzo e condivisione.",
    "sai usare telegraph": "S√¨, posso pubblicare contenuti su Telegraph! Se vuoi che pubblichi qualcosa, dimmi semplicemente 'Scrivimi un saggio su [argomento] e pubblicalo su Telegraph' e lo far√≤ per te!",
    "puoi pubblicare su telegraph": "Certamente! Posso generare contenuti e pubblicarli su Telegraph. Prova a chiedermi: 'Scrivimi un saggio su Roma e pubblicalo su Telegraph'",
    "come usare telegraph": "Per usare Telegraph con me, basta che mi chiedi di scrivere qualcosa e di pubblicarlo su Telegraph. Ad esempio: 'Scrivimi un articolo sulla storia di Roma e pubblicalo su Telegraph'",
    "cos'√® CES": "CES √® l'acronimo di Cogito Ergo Sum, un modello di intelligenza artificiale openspurce sviluppato da Mirko Yuri Donato. Attualmente sono disponibili due versioni: CES 1.0 (Cohere) e CES 1.5 (Gemini).",
    "cos'√® la modalit√† sperimentale": "La modalit√† sperimentale √® una funzionalit√† di ArcadiaAI che consente di testare nuove funzionalit√† e miglioramenti prima del rilascio ufficiale. Pu√≤ includere nuove capacit√†, modelli o strumenti.",
    "cos'√® la modalit√† sviluppatore": "La modalit√† sviluppatore √® una funzionalit√† di ArcadiaAI che consente agli sviluppatori di testare e implementare nuove funzionalit√†, modelli o strumenti. √à progettata per facilitare lo sviluppo e il miglioramento continuo del chatbot.",
    "che differenza c'√® tra la modalit√† sperimentale e la modalit√† sviluppatore": "La modalit√† sperimentale √® destinata agli utenti finali per testare nuove funzionalit√†, mentre la modalit√† sviluppatore √® per gli sviluppatori che vogliono implementare e testare nuove funzionalit√† o modelli. Entrambe le modalit√† possono coesistere e migliorare l'esperienza utente.",
    "cos'√® CES Plus": "CES Plus √® una versione avanzata di CES, ottimizzata nei ragionamenti e nella generazione di contenuti",
    "cos'√® CES 1.0": "CES 1.0 √® la prima versione del modello CES, sviluppato da Mirko Yuri Donato. Utilizza la tecnologia Cohere per generare contenuti e rispondere a domande. Tieni presente che questa versione verr√† dismessa a partire dal 20 Maggio 2025.",
    "cos'√® CES 1.5": "CES 1.5 √® la versione pi√π recente del modello CES, sviluppato da Mirko Yuri Donato. Utilizza la tecnologia Gemini per generare contenuti e rispondere a domande. Questa versione offre prestazioni migliorate rispetto a CES 1.0 ma inferiori a CES Plus",
    "come attivo la modalit√† sperimentale": "Per attivare la modalit√† sperimentale, basta chiedere a ArcadiaAI di attivarla usando il comando \"@impostazioni modalit√† sperimentale attiva\". Una volta attivata, potrai testare nuove funzionalit√† e miglioramenti.",
    "come attivo la modalit√† sviluppatore": "Per attivare la modalit√† sviluppatore, basta chiedere a ArcadiaAI di attivarla usando il comando \"@impostazioni modalit√† sviluppatore attiva\". Una volta attivata, potrai testare e implementare nuove funzionalit√† e modelli.",
    "come disattivo la modalit√† sperimentale": "Per disattivare la modalit√† sperimentale, basta chiedere a ArcadiaAI di disattivarla usando il comando \"@impostazioni modalit√† sperimentale disattiva\". Una volta disattivata, non potrai pi√π testare le nuove funzionalit√†.",
    "codice sorgente arcadiaai": "Il codice sorgente di ArcadiaAI √® pubblico! Puoi trovarlo con il comando @codice_sorgente oppure visitando la repository: https://github.com/Mirko-linux/Nova-Surf/tree/main/ArcadiaAI",
    "sai cercare su internet": "S√¨, posso cercare informazioni su Internet. Se hai bisogno di qualcosa in particolare dimmi @cerca e il termine di ricerca e io lo far√≤ per te",
    "sai usare google": "No, non posso usare Google, perch√© sono progrmmato per cercare solamente su DuckDuckGo. Posso cercare informazioni su Internet usando DuckDuckGo. Se hai bisogno di qualcosa in particolare dimmi @cerca e il termine di ricerca e io lo far√≤ per te",
    "come vengono salvate le conversazioni": "Le conversazioni vengoono salvate in locale sul tuo browser. Non vengono memorizzate su server esterni e non vengono condivise con terze parti. La tua privacy √® importante per noi.",
    "come posso cancellare le conversazioni": "Puoi cancellare le conversazioni andando nelle impostazioni del tuo browser e cancellando la cache e i cookie. In alternativa, puoi usare il comando @cancella_conversazione per eliminare la cronologia delle chat.",
    "cosa sono i cookie": "I cookie sono piccoli file di testo che i siti web memorizzano sul tuo computer per ricordare informazioni sulle tue visite. Possono essere utilizzati per tenere traccia delle tue preferenze, autenticarti e migliorare l'esperienza utente.",
}

# Trigger per le risposte predefinite
trigger_phrases = {
    "chi sei": ["chi sei", "chi sei tu", "tu chi sei", "presentati", "come ti chiami", "qual √® il tuo nome"],
    "cosa sai fare": ["cosa sai fare", "cosa puoi fare", "funzionalit√†", "capacit√†", "a cosa servi", "in cosa puoi aiutarmi"],
    "chi √® tobia testa": ["chi √® tobia testa", "informazioni su tobia testa", "parlami di tobia testa", "chi √® tobia teseo"],
    "chi √® mirko yuri donato": ["chi √® mirko yuri donato", "informazioni su mirko yuri donato", "parlami di mirko yuri donato", "chi ha creato arcadiaai"],
    "chi √® il presidente di arcadia": ["chi √® il presidente di arcadia", "presidente di arcadia", "chi guida arcadia", "capo di arcadia"],
    "chi √® il presidente di lumenaria": ["chi √® il presidente di lumenaria", "presidente di lumenaria", "chi guida lumenaria", "capo di lumenaria", "carlo cesare orlando presidente"],
    "cos'√® nova surf": ["cos'√® nova surf", "che cos'√® nova surf", "parlami di nova surf", "a cosa serve nova surf"],
    "chi ti ha creato": ["chi ti ha creato", "chi ti ha fatto", "da chi sei stato creato", "creatore di arcadiaai"],
    "chi √® ciua grazisky": ["chi √® ciua grazisky", "informazioni su ciua grazisky", "parlami di ciua grazisky"],
    "chi √® carlo cesare orlando": ["chi √® carlo cesare orlando", "informazioni su carlo cesare orlando", "parlami di carlo cesare orlando", "chi √® davide leone"],
    "chi √® omar lanfredi": ["chi √® omar lanfredi", "informazioni su omar lanfredi", "parlami di omar lanfredi"],
    "cos'√® arcadiaai": ["cos'√® arcadiaai", "che cos'√® arcadiaai", "parlami di arcadiaai", "a cosa serve arcadiaai"],
    "sotto che licenza √® distribuito arcadiaa": ["sotto che licenza √® distribuito arcadiaa", "licenza arcadiaai", "che licenza usa arcadiaai", "arcadiaai licenza"],
    "cosa sono le micronazioni": ["cosa sono le micronazioni", "micronazioni", "che cosa sono le micronazioni", "parlami delle micronazioni"],
    "cos'√® la repubblica di arcadia": ["cos'√® la repubblica di arcadia", "repubblica di arcadia", "che cos'√® la repubblica di arcadia", "parlami della repubblica di arcadia", "arcadia micronazione"],
    "cos'√® la repubblica di lumenaria": ["cos'√® la repubblica di lumenaria", "repubblica di lumenaria", "che cos'√® la repubblica di lumenaria", "parlami della repubblica di lumenaria", "lumenaria micronazione"],
    "chi √® salvatore giordano": ["chi √® salvatore giordano", "informazioni su salvatore giordano", "parlami di salvatore giordano"],
    "da dove deriva il nome arcadia": ["da dove deriva il nome arcadia", "origine nome arcadia", "significato nome arcadia", "perch√© si chiama arcadia"],
    "da dove deriva il nome lumenaria": ["da dove deriva il nome lumenaria", "origine nome lumenaria", "significato nome lumenaria", "perch√© si chiama lumenaria"],
    "da dove deriva il nome leonia": ["da dove deriva il nome leonia", "origine nome leonia", "significato nome leonia", "perch√© si chiama leonia"],
    "cosa si intende per open source": ["cosa si intende per open source", "open source significato", "che significa open source", "definizione di open source"],
    "arcadiaai √® un software libero": ["arcadiaai √® un software libero", "arcadiaai software libero", "arcadiaai √® libero", "software libero arcadiaai"],
    "cos'√® un chatbot": ["cos'√® un chatbot", "chatbot significato", "che significa chatbot", "definizione di chatbot"],
    "sotto che licenza sei distribuita": ["sotto che licenza sei distribuita", "licenza di arcadiaai", "che licenza usi", "arcadiaai licenza"],
    "sai usare telegraph": ["sai pubblicare su telegraph", "funzioni su telegraph", "hai telegraph integrato", "telegraph", "puoi usare telegraph"],
    "puoi pubblicare su telegraph": ["puoi pubblicare su telegraph", "pubblicare su telegraph", "supporti telegraph"],
    "come usare telegraph": ["come usare telegraph", "come funziona telegraph", "istruzioni telegraph"],
    "cos'√® CES" : ["cos √® CES", "CES", "che cos'√® CES", "definizione di CES"],
    "cos'√® la modalit√† sperimentale": ["cos'√® la modalit√† sperimentale", " parlami della modalit√† sperimentale", "che cos'√® la modalit√† sperimentale"],
    "cos'√® la modalit√† sviluppatore": ["cos'√® la modalit√† sviluppatore", " parlami della modalit√† sviluppatore", "che cos'√® la modalit√† sviluppatore"],
    "che differenza c'√® tra la modalit√† sperimentale e la modalit√† sviluppatore": ["che differenza c'√® tra la modalit√† sperimentale e la modalit√† sviluppatore", "differenza tra modalit√† sperimentale e sviluppatore", "modalit√† sperimentale vs sviluppatore"],
    "cos'√® CES Plus": ["cos'√® CES Plus", "che cazzo √® CES Plus", "che cos'√® CES Plus", "definizione di CES Plus"],
    "cos'√® CES 1.0": ["cos'√® CES 1.0", "che cos'√® CES 1.0", "definizione di CES 1.0"],
    "cos'√® CES 1.5": ["cos'√® CES 1.5", "che cos'√® CES 1.5", "definizione di CES 1.5"],
    "come attivo la modalit√† sperimentale": ["come attivo la modalit√† sperimentale", "attivare modalit√† sperimentale", "come usare la modalit√† sperimentale"],
    "come attivo la modalit√† sviluppatore": ["come attivo la modalit√† sviluppatore", "attivare modalit√† sviluppatore", "come usare la modalit√† sviluppatore"],
    "come disattivo la modalit√† sperimentale": ["come disattivo la modalit√† sperimentale", "disattivare modalit√† sperimentale", "come usare la modalit√† sperimentale"],
    "come disattivo la modalit√† sviluppatore": ["come disattivo la modalit√† sviluppatore", "disattivare modalit√† sviluppatore", "come usare la modalit√† sviluppatore"],
    "codice sorgente arcadiaai": [
    "dove posso trovare il codice sorgente di arcadiaai",
    "codice sorgente arcadiaai",
    "dove trovo il codice sorgente di arcadiaai",
    "dove si trova il codice sorgente di arcadiaai",
    "come posso vedere il codice sorgente di arcadiaai",
    "codice sorgente di arcadiaai",
    "dove posso trovare il codice sorgente tuo",
    "dove trovo il codice sorgente tuo"],
    "sai cercare su internet": ["sai cercare su internet", "cerca su internet", "puoi cercare su internet", "cerca"],
    "sai usare google": ["sai usare google", "puoi usare google", "cerca su google", "cerca su google per me"],
    "come vengono salvate le conversazioni": ["come vengono salvate le conversazioni", "dove vengono salvate le conversazioni", "salvataggio conversazioni", "come vengono salvate le chat"],
    "come posso cancellare le conversazioni": ["come posso cancellare le conversazioni", "cancellare conversazioni", "cancellare chat", "come cancellare le chat"],
}

import zipfile
import tempfile
import importlib.util
import json
def load_nsk_extensions():
    ext_dir = os.path.join(os.path.dirname(__file__), "extensions")
    print("DEBUG: Contenuto cartella extensions:", os.listdir(ext_dir))
    if not os.path.exists(ext_dir):
        os.makedirs(ext_dir)
    for filename in os.listdir(ext_dir):
        print("DEBUG: Analizzo file:", filename)
        if filename.endswith(".nsk") or filename.endswith(".zip"):
            nsk_path = os.path.join(ext_dir, filename)
            print("DEBUG: Provo ad aprire:", nsk_path)
            try:
                with zipfile.ZipFile(nsk_path, 'r') as zip_ref:
                    print("DEBUG: File zip aperto:", nsk_path)
                    with tempfile.TemporaryDirectory() as temp_dir:
                        zip_ref.extractall(temp_dir)
                        print("DEBUG: Estratti file:", os.listdir(temp_dir))
                        manifest_path = os.path.join(temp_dir, "manifest.json")
                        print("DEBUG: Cerco manifest:", manifest_path, "esiste?", os.path.exists(manifest_path))
                        if not os.path.exists(manifest_path):
                            print(f"Manifest mancante in {filename}")
                            continue
                        with open(manifest_path, "r", encoding="utf-8") as f:
                            manifest = json.load(f)
                        entrypoint = manifest.get("entrypoint", "main.py")
                        entry_path = os.path.join(temp_dir, entrypoint)
                        print("DEBUG: Cerco entrypoint:", entry_path, "esiste?", os.path.exists(entry_path))
                        if not os.path.exists(entry_path):
                            print(f"Entrypoint mancante in {filename}")
                            continue
                        module_name = f"nsk_{filename.replace('.nsk','')}"
                        spec = importlib.util.spec_from_file_location(module_name, entry_path)
                        module = importlib.util.module_from_spec(spec)
                        try:
                            spec.loader.exec_module(module)
                            if hasattr(module, "can_handle") and hasattr(module, "handle"):
                                EXTENSIONS[module_name] = module
                                print(f"Estensione caricata: {module_name}")
                            else:
                                print(f"Estensione {module_name} non valida (manca can_handle o handle)")
                        except Exception as e:
                            print(f"Errore caricamento estensione {filename}: {e}")
            except Exception as e:
                print(f"Errore apertura zip {filename}: {e}")
    print("Estensioni caricate:", list(EXTENSIONS.keys()))
load_nsk_extensions()

# Funzione per pubblicare su Telegraph
def publish_to_telegraph(title, content):
    """Pubblica contenuti su Telegraph."""
    url = "https://api.telegra.ph/createPage"
    headers = {"Content-Type": "application/json"}
    
    paragraphs = [p.strip() for p in content.split('\n') if p.strip()]
    content_formatted = [{"tag": "p", "children": [p]} for p in paragraphs[:50]]
    
    payload = {
        "access_token": TELEGRAPH_API_KEY,
        "title": title[:256],
        "content": content_formatted,
        "author_name": "ArcadiaAI"
    }
    
    try:
        res = requests.post(url, headers=headers, json=payload, timeout=15)
        res.raise_for_status()
        result = res.json()
        if result.get("ok"):
            return result.get("result", {}).get("url", "‚ö†Ô∏è URL non disponibile")
        return "‚ö†Ô∏è Pubblicazione fallita"
    except requests.exceptions.RequestException as e:
        print(f"Errore pubblicazione Telegraph (connessione): {str(e)}")
        return f"‚ö†Ô∏è Errore di connessione a Telegraph: {str(e)}"
    except Exception as e:
        print(f"Errore pubblicazione Telegraph: {str(e)}")
        return f"‚ö†Ô∏è Errore durante la pubblicazione: {str(e)}"

# Funzione per generare contenuti con Gemini (CES 1.5)
def generate_with_gemini(prompt, title):
    """Genera contenuti con Gemini e pubblica su Telegraph."""
    if not gemini_model:
        return None, "‚ùå Gemini (CES 1.5) non √® configurato"
    
    try:
        # Aggiungi contesto identitario
        full_prompt = (
            "Sei ArcadiaAI, un chatbot open source creato da Mirko Yuri Donato. "
            "Stai generando un contenuto che verr√† pubblicato su Telegraph. "
            "Il contenuto deve essere accurato, ben strutturato e mantenere "
            "uno stile professionale. Ecco la richiesta:\n\n"
            f"{prompt}"
        )
        
        response = gemini_model.generate_content(
            full_prompt,
            generation_config={
                "max_output_tokens": 3000,
                "temperature": 0.8
            }
        )
        
        if not response.text:
            return None, "‚ùå Impossibile generare il contenuto"
        
        telegraph_url = publish_to_telegraph(title, response.text)
        return response.text, telegraph_url
    
    except Exception as e:
        print(f"Errore generazione contenuto Gemini: {str(e)}")
        return None, f"‚ùå Errore durante la generazione: {str(e)}"    


def extract_article_content(soup):
    article_body = soup.find('article') or soup.find('div', class_='article-body')
    if article_body:
        return article_body
    return None

def extract_main_content(soup):
    main_content = soup.find('main') or soup.find('div', id='content') or soup.find('div', class_='main-content')
    if main_content:
        return main_content
    return soup.body

def search_web(query, lang="it-IT"):
    try:
        ddg_results = search_duckduckgo(query, lang)

        if not ddg_results or len(ddg_results) < 3:
            print("DEBUG: Pochi risultati da DDG, tentativo con Brave.")
            brave_results = search_brave(query, lang)
            results = (ddg_results or []) + (brave_results or [])
        else:
            results = ddg_results

        filtered = filter_results(results, query)
        verified = verify_results(filtered)

        return verified[:3]
    except Exception as e:
        print(f"Errore generale nella funzione search_web: {str(e)}")
        return []

def search_duckduckgo(query, lang):
    url = f"https://html.duckduckgo.com/html/?q={urllib.parse.quote(query)}&kl={lang[:2]}-{lang[-2:]}"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Accept-Language": lang
    }

    try:
        res = requests.get(url, headers=headers, timeout=12)
        res.raise_for_status()

        print("DEBUG HTML DDG (primi 1000 caratteri):", res.text[:1000])

        soup = BeautifulSoup(res.text, 'html.parser')
        results = []

        for result_div in soup.select('div.result, div.web-result'):
            link_tag = result_div.find('a', class_='result__a')
            if link_tag and 'href' in link_tag.attrs:
                href = extract_real_url(link_tag['href'])
                if href:
                    title = link_tag.get_text(strip=True)

                    snippet_tag = result_div.find('div', class_='result__body')
                    if not snippet_tag:
                        snippet_tag = result_div.find('a', class_='result__snippet')

                    snippet_text = snippet_tag.get_text(' ', strip=True) if snippet_tag else ""

                    if title or snippet_text:
                        results.append({
                            'url': href,
                            'title': title,
                            'snippet': snippet_text,
                            'source': 'duckduckgo'
                        })
        print(f"DEBUG: Trovati {len(results)} risultati da DuckDuckGo.")
        return results[:5]
    except requests.exceptions.RequestException as e:
        print(f"Errore nella richiesta HTTP a DuckDuckGo: {e}")
        return []
    except Exception as e:
        print(f"Errore generico in search_duckduckgo: {e}")
        return []

def search_brave(query, lang):
    url = f"https://search.brave.com/search?q={urllib.parse.quote(query)}&lr=lang_{lang[:2]}"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Accept-Language": lang
    }

    try:
        res = requests.get(url, headers=headers, timeout=12)
        res.raise_for_status()

        print("DEBUG HTML Brave (primi 1000 caratteri):", res.text[:1000])

        soup = BeautifulSoup(res.text, 'html.parser')
        results = []

        for result_div in soup.select('.result'):
            link_tag = result_div.find('a')
            if link_tag and link_tag.has_attr('href'):
                href = extract_real_url(link_tag['href'])
                if href:
                    title = link_tag.get_text(strip=True)
                    snippet_tag = result_div.find(class_='snippet-content')
                    snippet_text = snippet_tag.get_text(' ', strip=True) if snippet_tag else ""

                    if title or snippet_text:
                        results.append({
                            'url': href,
                            'title': title,
                            'snippet': snippet_text,
                            'source': 'brave'
                        })
        print(f"DEBUG: Trovati {len(results)} risultati da Brave.")
        return results[:5]
    except requests.exceptions.RequestException as e:
        print(f"Errore nella richiesta HTTP a Brave: {e}")
        return []
    except Exception as e:
        print(f"Errore generico in search_brave: {str(e)}")
        return []

def extract_real_url(href):
    if href.startswith('http'):
        return href
    elif href.startswith('//duckduckgo.com/l/?uddg='):
        parsed = urllib.parse.urlparse('https:' + href)
        query = urllib.parse.parse_qs(parsed.query)
        return urllib.parse.unquote(query.get('uddg', [''])[0]) or None
    return None

def filter_results(results, query):
    if not results:
        return []

    unique = {r['url']: r for r in results}.values()

    query_words = set(query.lower().split())
    ranked = []

    for result in unique:
        title = result.get('title', '').lower()
        snippet = result.get('snippet', '').lower()

        title_score = sum(1 for word in query_words if word in title)
        snippet_score = sum(1 for word in query_words if word in snippet)

        domain = urllib.parse.urlparse(result.get('url', '')).netloc
        trust_score = 1.0
        if any(bad_word in domain for bad_word in ['porn', 'adult', 'spam', 'scam']):
            trust_score = 0.1

        total_score = (title_score * 2 + snippet_score) * trust_score
        ranked.append((total_score, result))

    ranked.sort(reverse=True, key=lambda x: x[0])
    return [r[1] for r in ranked]

def verify_results(results):
    verified = []

    for result in results[:5]:
        try:
            head = requests.head(result['url'], timeout=5, allow_redirects=True)
            if head.status_code == 200:
                result['verified_at'] = datetime.now().isoformat()
                verified.append(result)
            else:
                print(f"DEBUG: URL non accessibile ({head.status_code}): {result['url']}")
        except requests.exceptions.RequestException as e:
            print(f"DEBUG: Errore di richiesta HEAD per {result['url']}: {e}")
        except Exception as e:
            print(f"DEBUG: Errore generico durante la verifica per {result['url']}: {e}")

    return verified or results[:3]

def extract_content(url):
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Accept-Language": "it-IT,it;q=0.9"
    }

    try:
        res = requests.get(url, headers=headers, timeout=20)
        res.raise_for_status()

        if 'text/html' not in res.headers.get('Content-Type', ''):
            print(f"DEBUG: URL {url} non √® HTML, Content-Type: {res.headers.get('Content-Type')}")
            return None

        soup = BeautifulSoup(res.text, 'lxml')

        for element in soup(['script', 'style', 'nav', 'footer', 'iframe', 'header', 'aside', 'form', 'button', 'noscript']):
            element.decompose()

        content = extract_article_content(soup) or extract_main_content(soup)

        if content:
            text = ' '.join(content.get_text(' ', strip=True).split())
            text = re.sub(r'\s+', ' ', text)
            return text[:2000]
        
        print(f"DEBUG: Nessun contenuto principale estratto da {url}")
        return None

    except requests.exceptions.RequestException as e:
        print(f"Errore di richiesta HTTP durante l'estrazione da {url}: {e}")
        return None
    except Exception as e:
        print(f"Errore generico durante l'estrazione da {url}: {str(e)}")
        return None

def estrai_testo_da_url(url):
    try:
        res = requests.get(url, timeout=10)
        res.raise_for_status()
        soup = BeautifulSoup(res.text, "html.parser")
        paragraphs = soup.find_all("p")
        testo = " ".join(p.get_text(strip=True) for p in paragraphs)
        return testo.strip()
    except requests.exceptions.RequestException as e:
        print(f"Errore di richiesta HTTP durante l'estrazione del testo da {url}: {e}")
        return ""
    except Exception as e:
        print(f"Errore generico durante l'estrazione del testo da {url}: {e}")
        return ""
    
def extract_article_content(soup):
    """Cerca contenuto in tag article/section"""
    for tag in ['article', 'main', 'section', 'div.article', 'div.content']:
        element = soup.select_one(tag) if '.' in tag else soup.find(tag)
        if element and len(element.get_text(strip=True)) > 100:
            return element
    return None

def extract_main_content(soup):
    """Estrae il contenuto principale con euristica"""
    paragraphs = soup.find_all('p')
    if len(paragraphs) > 3:
        container = paragraphs[0].parent
        if container and len(container.get_text(strip=True)) > 200:
            return container
    return None

import requests

def is_valid_audio(url):
    try:
        response = requests.head(url, timeout=5)
        content_length = int(response.headers.get("Content-Length", 0))
        return content_length > 1000  # ad esempio, almeno 1 KB
    except:
        return False
    

import requests
from uuid import uuid4
# Endpoint per generare musica

@app.route("/api/ces-image", methods=["POST"])
def ces_image():
    try:
        data = request.get_json()
        prompt = data.get("prompt", "").strip() if data else ""
        print(f"[CES-IMAGE] Prompt ricevuto: {prompt}")

        if not prompt:
            return jsonify({"error": "Prompt mancante o vuoto."}), 400

        # üõ°Ô∏è Filtro anti-abusi: parole non consentite
        PAROLE_BANNATE = [
            "nudo", "nudit√†", "naked", "porn", "porno", "pornografico", "sessuale",
            "sex", "sesso", "genitali", "masturb", "boobs", "dildo", "nsfw", "erotico", 
            "fetish", "xxx", "pene", "vagina", "anale", "seni", "hot", "orgasmo"
        ]

        if any(parola in prompt.lower() for parola in PAROLE_BANNATE):
            print(f"[CES-IMAGE] ‚ùå Prompt rifiutato per contenuto vietato: {prompt}")
            return jsonify({"error": "‚ùå Questo prompt non √® consentito. Per favore, rispetta le linee guida."}), 403

        # Codifica il prompt per l'URL
        encoded_prompt = urllib.parse.quote(prompt)
        image_url = f"https://image.pollinations.ai/prompt/{encoded_prompt}"

        return jsonify({"image_url": image_url})

    except Exception as e:
        print(f"[CES-IMAGE] Errore interno: {e}")
        return jsonify({"error": f"Errore interno: {str(e)}"}), 500

from flask import Flask, Response
@app.route('/')
def home():
    html = """
<!DOCTYPE html>
<html lang="it">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>ArcadiaAI Chat</title>
    <link rel="stylesheet" href="/static/style.css">
    <style>
        /* Migliora layout su schermi piccoli */
        @media (max-width: 768px) {
            body {
                display: flex;
                flex-direction: column;
                margin: 0;
                padding: 0;
                overflow-x: hidden;
            }

            #sidebar {
                width: 100%;
                position: static;
                padding: 10px;
                box-sizing: border-box;
                border-bottom: 1px solid #ccc;
            }

            #chat-area {
                width: 100%;
                padding: 10px;
                box-sizing: border-box;
            }

            #chatbox {
                max-height: 50vh;
                overflow-y: auto;
                margin-bottom: 10px;
            }

            #input-area {
                display: flex;
                flex-direction: column;
                gap: 10px;
            }

            #input {
                width: 100%;
                font-size: 16px;
                padding: 10px;
            }

            #attach-btn,
            #send-btn {
                width: 100%;
                padding: 10px;
                font-size: 16px;
            }

            #canvas-overlay > div {
                width: 95% !important;
                max-width: 100% !important;
                height: auto;
                margin: 20px auto;
            }

            #canvas-toolbar {
                flex-direction: column;
                gap: 5px;
            }

            #canvas-editor, #canvas-properties {
                flex: 1 1 100%;
                margin-bottom: 20px;
            }

            #canvas-content {
                height: 300px;
            }

            .canvas-element {
                max-width: 90%;
            }
        }

        /* Stili base */
        * {
            box-sizing: border-box;
            margin: 0;
            padding: 0;
        }

        body {
            display: flex;
            height: 100vh;
            background-color: #343541;
            color: #fff;
            font-family: 'Segoe UI', Roboto, -apple-system, BlinkMacSystemFont, sans-serif;
            line-height: 1.5;
        }

        /* Sidebar */
        #sidebar {
            width: 260px;
            background: #202123;
            padding: 15px;
            overflow-y: auto;
            border-right: 1px solid #444;
            display: flex;
            flex-direction: column;
        }

        #sidebar h2 {
            color: #fff;
            margin: 10px 0 20px;
            text-align: center;
            font-size: 1.4rem;
            font-weight: 600;
        }

        #sidebar button {
            width: 100%;
            padding: 10px;
            margin-bottom: 10px;
            border: none;
            border-radius: 5px;
            background: #10a37f;
            color: white;
            cursor: pointer;
            font-weight: 500;
            transition: background-color 0.2s ease;
        }

        #sidebar button:hover {
            background: #0e8e6d;
        }

        #sidebar button:active {
            transform: scale(0.98);
        }

        /* Lista chat */
        #chat-list {
            list-style: none;
            margin-top: 15px;
            flex-grow: 1;
            overflow-y: auto;
        }

        #chat-list li {
            padding: 12px;
            margin-bottom: 8px;
            background: #353740;
            border-radius: 5px;
            cursor: pointer;
            color: #ececf1;
            position: relative;
            transition: background-color 0.2s ease;
            overflow: hidden;
            text-overflow: ellipsis;
            white-space: nowrap;
        }

        #chat-list li:hover {
            background: #40414f;
        }

        #chat-list li.active {
            background: #4e4f5c;
        }

        #chat-list li.empty {
            text-align: center;
            cursor: default;
            background: transparent;
            color: #a1a1a9;
        }

        /* Area chat */
        #chat-area {
            flex: 1;
            display: flex;
            flex-direction: column;
            position: relative;
        }

        #chatbox {
            flex: 1;
            padding: 20px;
            overflow-y: auto;
            background: #343541;
            scroll-behavior: smooth;
        }

        /* Messaggi */
        .message {
            margin: 12px 0;
            padding: 12px 16px;
            border-radius: 8px;
            max-width: 85%;
            line-height: 1.5;
            word-wrap: break-word;
            animation: fadeIn 0.3s ease;
        }

        @keyframes fadeIn {
            from { opacity: 0; transform: translateY(10px); }
            to { opacity: 1; transform: translateY(0); }
        }

        .user-message {
            background: #10a37f;
            color: white;
            margin-left: auto;
            border-bottom-right-radius: 0;
        }

        .ai-message {
            background: #444654;
            color: white;
            margin-right: auto;
            border-bottom-left-radius: 0;
        }

        /* Input area */
        #input-area {
            display: flex;
            padding: 15px;
            background: #40414f;
            position: sticky;
            bottom: 0;
            gap: 10px;
        }

        #input {
            flex: 1;
            padding: 12px;
            font-size: 1rem;
            border: none;
            border-radius: 6px;
            background: #53545f;
            color: white;
            transition: outline 0.2s ease;
        }

        #input:focus {
            outline: 2px solid #10a37f;
            background: #5e5f6b;
        }

        #input::placeholder {
            color: #a1a1a9;
        }

        /* Pulsanti */
        button {
            background: #10a37f;
            color: white;
            border: none;
            padding: 0 20px;
            border-radius: 6px;
            cursor: pointer;
            font-weight: 500;
            transition: all 0.2s ease;
            min-width: 80px;
            height: 42px;
        }

        button:hover {
            background: #0e8e6d;
        }

        button:active {
            transform: scale(0.98);
        }

        button:disabled {
            background: #5f6d6a;
            cursor: not-allowed;
        }

        /* Menu contestuale */
        .chat-menu {
            display: none;
            position: absolute;
            background: #2d2e35;
            border-radius: 5px;
            z-index: 100;
            box-shadow: 0 2px 10px rgba(0,0,0,0.2);
            min-width: 150px;
            border: 1px solid #444;
        }

        .chat-menu div {
            padding: 10px 15px;
            cursor: pointer;
            transition: background-color 0.2s ease;
            color: #ececf1;
            font-size: 14px;
        }

        .chat-menu div:hover {
            background: #40414f;
        }

        .chat-title-container {
            display: flex;
            justify-content: space-between;
            align-items: center;
            width: 100%;
        }

        .chat-actions {
            visibility: hidden;
            cursor: pointer;
            padding: 0 5px;
            color: #a1a1a9;
            font-size: 18px;
            user-select: none;
        }

        li:hover .chat-actions {
            visibility: visible;
        }

        /* Modali */
        .modal {
            display: none;
            position: fixed;
            z-index: 1001;
            left: 0;
            top: 0;
            width: 100%;
            height: 100%;
            background-color: rgba(0, 0, 0, 0.5);
        }

        .modal-content {
            background-color: #40414f;
            margin: 15% auto;
            padding: 20px;
            border: 1px solid #555;
            width: 300px;
            border-radius: 8px;
            position: relative;
            color: #ececf1;
        }

        .modal-content h3 {
            margin-bottom: 15px;
            color: #fff;
        }

        .modal-content p {
            margin-bottom: 15px;
            color: #a1a1a9;
        }

        .close {
            position: absolute;
            right: 15px;
            top: 10px;
            font-size: 24px;
            cursor: pointer;
            color: #a1a1a9;
        }

        .close:hover {
            color: #fff;
        }

        /* Opzioni esportazione */
        .export-options {
            display: flex;
            justify-content: space-between;
            margin-top: 20px;
            gap: 10px;
        }

        .export-btn {
            padding: 10px;
            background: #10a37f;
            color: white;
            border: none;
            border-radius: 5px;
            cursor: pointer;
            flex: 1;
            text-align: center;
            font-size: 14px;
        }

        .export-btn:hover {
            background: #0e8e6d;
        }

        /* Input password */
        input[type="password"] {
            width: 100%;
            padding: 10px;
            margin: 15px 0;
            box-sizing: border-box;
            border: 1px solid #555;
            border-radius: 4px;
            background: #53545f;
            color: #fff;
        }

        /* Messaggio bloccato */
        .locked-message {
            text-align: center;
            padding: 40px 20px;
            color: #a1a1a9;
            height: 100%;
            display: flex;
            flex-direction: column;
            justify-content: center;
            align-items: center;
        }

        .locked-message p {
            margin-bottom: 20px;
            font-size: 16px;
        }

        #unlock-btn {
            padding: 10px 20px;
            background: #10a37f;
            color: white;
            border: none;
            border-radius: 5px;
            cursor: pointer;
        }

        #unlock-btn:hover {
            background: #0e8e6d;
        }

        /* Allegati */
        #attach-btn {
            background: none;
            border: none;
            font-size: 1.5em;
            cursor: pointer;
            padding: 0 10px;
            color: #a1a1a9;
        }

        #attach-btn:hover {
            color: #fff;
        }

        #attachments-preview {
            display: flex;
            flex-wrap: wrap;
            gap: 8px;
            margin-bottom: 10px;
        }

        .attachment-preview {
            border: 1px solid #555;
            border-radius: 4px;
            padding: 5px;
            display: flex;
            align-items: center;
            background: #53545f;
            max-width: 200px;
        }

        .attachment-preview img {
            max-height: 50px;
            max-width: 50px;
            margin-right: 8px;
            border-radius: 3px;
        }

        .file-icon {
            margin-right: 8px;
            font-size: 1.2em;
        }

        .remove-attachment {
            background: none;
            border: none;
            color: #a1a1a9;
            cursor: pointer;
            margin-left: auto;
            padding: 0 5px;
        }

        .remove-attachment:hover {
            color: #f00;
        }

        .message-attachments {
            margin-bottom: 8px;
        }

        .message-attachments img {
            max-width: 100%;
            max-height: 200px;
            border-radius: 4px;
        }

        .document-attachment {
            display: block;
            padding: 8px;
            background: #53545f;
            border-radius: 4px;
            color: #10a37f;
            text-decoration: none;
            margin-bottom: 5px;
            border: 1px solid #555;
        }

        .document-attachment:hover {
            text-decoration: underline;
            background: #5e5f6b;
        }

        /* Scrollbar */
        ::-webkit-scrollbar {
            width: 8px;
        }

        ::-webkit-scrollbar-track {
            background: #2d2d2d;
        }

        ::-webkit-scrollbar-thumb {
            background: #555;
            border-radius: 4px;
        }

        ::-webkit-scrollbar-thumb:hover {
            background: #666;
        }

        /* Responsive */
        @media (max-width: 768px) {
            #sidebar {
                width: 200px;
                padding: 10px;
            }

            .message {
                max-width: 90%;
                padding: 10px 14px;
            }
        }

        @media (max-width: 480px) {
            body {
                flex-direction: column;
            }

            #sidebar {
                width: 100%;
                height: auto;
                max-height: 200px;
                border-right: none;
                border-bottom: 1px solid #444;
            }

            #chat-area {
                height: calc(100vh - 200px);
            }

            .modal-content {
                width: 90%;
                margin: 50% auto;
            }
        }

        /* Toggle Ricerca Web */
        #web-search-toggle {
            padding: 10px 15px;
            background: #40414f;
            border-top: 1px solid #555;
            display: flex;
            align-items: center;
        }

        #web-search-toggle label {
            display: flex;
            align-items: center;
            gap: 8px;
            font-size: 14px;
            color: #a1a1a9;
            cursor: pointer;
        }

        #web-search-toggle input[type="checkbox"] {
            appearance: none;
            width: 18px;
            height: 18px;
            border: 1px solid #10a37f;
            border-radius: 4px;
            cursor: pointer;
            position: relative;
        }

        #web-search-toggle input[type="checkbox"]:checked {
            background: #10a37f;
        }

        #web-search-toggle input[type="checkbox"]:checked::after {
            content: "‚úì";
            position: absolute;
            color: white;
            font-size: 12px;
            top: 50%;
            left: 50%;
            transform: translate(-50%, -50%);
        }

        /* Stili per i modelli */
        #model-info {
            font-size: 0.8em;
            color: #666;
            margin-top: 5px;
        }

        .model-icon {
            margin-right: 8px;
            font-size: 1.1em;
        }

        .model-indicator {
            font-size: 0.7em;
            color: #666;
            margin-bottom: 5px;
        }

        .verify-btn {
            background-color: #f5f5f5;
            border: none;
            font-size: 16px;
            cursor: pointer;
            transition: 0.3s;
        }

        .verify-btn:hover {
            background-color: #dcefff;
        }

        .ces-zero-result {
            font-size: 14px;
            font-weight: bold;
        }
    </style>
</head>
<body>
    <div id="sidebar">
        <h2>üß† ArcadiaAI</h2>
        <div id="api-selection">
            <label for="api-provider">Modello:</label>
            <select id="api-provider">
                <option value="gemini">CES 1.5</option>
                <option value="cesplus">CES Plus</option>
                <option value="ces360">CES 360</option>
                <option value="deepseek">DeepSeek R1</option>
                <option value="llama">Llama 3.1</option>
            </select>
        </div>
        <button id="new-chat-btn">‚ûï Nuova Chat</button>
        <button id="canvas-toggle-btn">üé® Canvas</button>
        <button id="clear-chats-btn" style="margin-top: 10px;">üóëÔ∏è Elimina Tutto</button>
        <button id="settings-btn">‚öôÔ∏è Impostazioni</button>
        <button id="sound-forge">üîä Sound Forge</button>
        <div id="settings-panel" style="display: none; padding: 10px;">
            <label for="language-select">Lingua:</label>
            <select id="language-select">
                <option value="it">Italiano</option>
                <option value="en">English</option>
            </select>
            <br>
            <label for="theme-select">Tema:</label>
            <select id="theme-select">
                <option value="light">Chiaro</option>
                <option value="dark">Scuro</option>
            </select>
            <br>
            <label><input type="checkbox" id="dev-mode-toggle"> Modalit√† sviluppatore</label><br>
            <label><input type="checkbox" id="experimental-mode-toggle"> Modalit√† sperimentale</label>
        </div>
        <ul id="chat-list"></ul>
    </div>
    <div id="chat-area">
        <div id="status-message" style="display: none; padding: 5px; background-color: #e0f7fa; border-radius: 5px; margin-bottom: 10px; text-align: center; color: #00796b;">
        </div>
        <div id="chatbox"></div>
        <div id="input-area">
            <input id="input" type="text" placeholder="Scrivi un messaggio..." autocomplete="off" aria-label="Messaggio">
            <input type="file" id="file-input" style="display:none" multiple aria-label="Allega file">
            <button id="attach-btn" title="Allega file" aria-label="Allega file">+</button>
            <button id="send-btn" aria-label="Invia messaggio">Invia</button>
        </div>
    </div>

    <!-- Canvas Overlay -->
    <div id="canvas-overlay" style="display:none; position:fixed; top:0; left:0; width:100%; height:100%; background:rgba(0,0,0,0.8); z-index:1000;">
        <div style="background:#fff; margin:20px auto; padding:20px; width:80%; max-width:900px; border-radius:10px; max-height:90vh; overflow:auto;">
            <div style="display:flex; justify-content:space-between; align-items:center; margin-bottom:15px;">
                <h2>üé® Canvas Editor</h2>
                <button id="close-canvas-btn" style="background:#ff4444; color:white; border:none; padding:5px 10px; border-radius:5px;">Chiudi</button>
            </div>
           
            <div id="canvas-toolbar" style="margin-bottom:15px; display:flex; gap:10px; flex-wrap:wrap;">
                <button class="canvas-tool" data-type="testo">‚úèÔ∏è Testo</button>
                <button class="canvas-tool" data-type="codice">{} Codice</button>
                <button class="canvas-tool" data-type="grafico">üìä Grafico</button>
                <button id="render-canvas-btn">üîÑ Anteprima</button>
                <select id="canvas-export-type" style="padding:5px;">
                    <option value="telegraph">Telegraph</option>
                    <option value="immagine">Immagine</option>
                    <option value="pdf">PDF</option>
                </select>
                <button id="export-canvas-btn">üì§ Esporta</button>
            </div>
           
            <div style="display:flex; gap:20px;">
                <div id="canvas-editor" style="flex:2; border:1px solid #ddd; padding:15px; min-height:500px; background:#f9f9f9;">
                    <div id="canvas-content" style="position:relative; width:100%; height:500px; background:white; border:1px dashed #ccc; overflow:hidden;">
                        <!-- Elementi canvas verranno aggiunti qui -->
                    </div>
                </div>
               
                <div id="canvas-properties" style="flex:1; border:1px solid #ddd; padding:15px;">
                    <h3>Propriet√† Elemento</h3>
                    <div id="element-properties-form" style="display:none;">
                        <div style="margin-bottom:10px;">
                            <label>Tipo:</label>
                            <input type="text" id="element-type" readonly style="width:100%; padding:5px;">
                        </div>
                        <div style="margin-bottom:10px;">
                            <label>Contenuto:</label>
                            <textarea id="element-content" style="width:100%; height:100px; padding:5px;"></textarea>
                        </div>
                        <div style="display:flex; gap:10px; margin-bottom:10px;">
                            <div>
                                <label>X:</label>
                                <input type="number" id="element-x" style="width:60px; padding:5px;">
                            </div>
                            <div>
                                <label>Y:</label>
                                <input type="number" id="element-y" style="width:60px; padding:5px;">
                            </div>
                        </div>
                        <button id="update-element-btn" style="background:#4CAF50; color:white; border:none; padding:5px 10px; border-radius:5px;">Aggiorna</button>
                        <button id="delete-element-btn" style="background:#ff4444; color:white; border:none; padding:5px 10px; border-radius:5px; margin-left:10px;">Elimina</button>
                    </div>
                    <div id="no-element-selected" style="color:#666; text-align:center; margin-top:50px;">
                        Seleziona un elemento per modificarlo
                    </div>
                </div>
            </div>
        </div>
    </div>

    <div id="message-display"></div>

    <script src="/static/script.js"></script>
    <script src="https://js.puter.com/v2/"></script>
    <script>
        // Inizializzazione Puter.js (esistente)
        document.addEventListener("DOMContentLoaded", async () => {
            if (typeof puter === "undefined") {
                alert("‚ùå Puter.js non √® stato caricato correttamente!");
                return;
            }

            console.log("‚úÖ Puter.js caricato con successo");
            cesAi = await puter.use("meta-llama/llama-3.3-70b-instruct");
            console.log("Modello attivo:", cesAi.model);
        });

        const soundForgeButton = document.getElementById('sound-forge');
        if (soundForgeButton) {
            soundForgeButton.addEventListener('click', () => {
                window.open('https://huggingface.co/spaces/mirkodonato08/arcadiaai-soundforge', '_blank');
            });
        } else {
            console.error("‚ùå Pulsante 'Sound Forge' non trovato! Assicurati che l'ID sia 'sound-forge'.");
        }

        // Gestione Canvas
        let currentCanvas = null;
        let selectedElement = null;

        // Toggle Canvas UI
        document.getElementById('canvas-toggle-btn').addEventListener('click', () => {
            document.getElementById('canvas-overlay').style.display = 'block';
            initCanvas();
        });

        document.getElementById('close-canvas-btn').addEventListener('click', () => {
            document.getElementById('canvas-overlay').style.display = 'none';
        });

        // Strumenti Canvas
        document.querySelectorAll('.canvas-tool').forEach(btn => {
            btn.addEventListener('click', function() {
                const type = this.dataset.type;
                addCanvasElement(type);
            });
        });

        function initCanvas() {
            const canvasContent = document.getElementById('canvas-content');
            canvasContent.innerHTML = '';
           
            if (!currentCanvas) {
                currentCanvas = {
                    id: 'canvas-' + Date.now(),
                    elements: []
                };
            }

            // Render elementi esistenti
            currentCanvas.elements.forEach(el => {
                renderCanvasElement(el);
            });
        }

        function addCanvasElement(type) {
            const newElement = {
                id: 'el-' + Date.now(),
                type: type,
                content: type === 'codice' ? '// Scrivi il tuo codice qui' :
                        type === 'grafico' ? 'Grafico generato' : 'Nuovo testo',
                position: { x: 50, y: 50 },
                style: {}
            };
           
            currentCanvas.elements.push(newElement);
            renderCanvasElement(newElement);
        }

        function renderCanvasElement(element) {
            const canvas = document.getElementById('canvas-content');
            const el = document.createElement('div');
            el.className = 'canvas-element';
            el.dataset.id = element.id;
           
            el.style.position = 'absolute';
            el.style.left = element.position.x + 'px';
            el.style.top = element.position.y + 'px';
            el.style.padding = '10px';
            el.style.background = '#fff';
            el.style.border = '1px solid #ddd';
            el.style.cursor = 'move';
            el.style.maxWidth = '300px';
           
            if (element.type === 'codice') {
                el.innerHTML = `<pre>${element.content}</pre>`;
                el.style.background = '#f5f5f5';
            } else if (element.type === 'grafico') {
                el.innerHTML = `<div class="graph-placeholder">${element.content}</div>`;
                el.style.textAlign = 'center';
            } else {
                el.textContent = element.content;
            }
           
            // Selezione elemento
            el.addEventListener('click', (e) => {
                e.stopPropagation();
                selectCanvasElement(element, el);
            });
           
            // Drag and drop
            el.addEventListener('mousedown', startDrag);
           
            canvas.appendChild(el);
        }

        function selectCanvasElement(element, htmlElement) {
            selectedElement = element;
           
            // Highlight elemento selezionato
            document.querySelectorAll('.canvas-element').forEach(el => {
                el.style.boxShadow = 'none';
            });
            htmlElement.style.boxShadow = '0 0 0 2px #4CAF50';
           
            // Popola form propriet√†
            document.getElementById('element-properties-form').style.display = 'block';
            document.getElementById('no-element-selected').style.display = 'none';
           
            document.getElementById('element-type').value = element.type;
            document.getElementById('element-content').value = element.content;
            document.getElementById('element-x').value = element.position.x;
            document.getElementById('element-y').value = element.position.y;
        }

        // Gestione drag and drop
        function startDrag(e) {
            const element = e.target.closest('.canvas-element');
            if (!element) return;
           
            const startX = e.clientX;
            const startY = e.clientY;
            const startLeft = parseInt(element.style.left) || 0;
            const startTop = parseInt(element.style.top) || 0;
           
            function moveHandler(e) {
                const dx = e.clientX - startX;
                const dy = e.clientY - startY;
               
                element.style.left = (startLeft + dx) + 'px';
                element.style.top = (startTop + dy) + 'px';
            }
           
            function upHandler() {
                document.removeEventListener('mousemove', moveHandler);
                document.removeEventListener('mouseup', upHandler);
               
                // Aggiorna posizione nell'oggetto canvas
                const elementId = element.dataset.id;
                const elementObj = currentCanvas.elements.find(el => el.id === elementId);
                if (elementObj) {
                    elementObj.position = {
                        x: parseInt(element.style.left),
                        y: parseInt(element.style.top)
                    };
                }
            }
           
            document.addEventListener('mousemove', moveHandler);
            document.addEventListener('mouseup', upHandler);
        }

        // Aggiornamento elementi
        document.getElementById('update-element-btn').addEventListener('click', () => {
            if (!selectedElement) return;
           
            selectedElement.content = document.getElementById('element-content').value;
            selectedElement.position = {
                x: parseInt(document.getElementById('element-x').value) || 0,
                y: parseInt(document.getElementById('element-y').value) || 0
            };
           
            // Re-render
            initCanvas();
        });

        // Elimina elemento
        document.getElementById('delete-element-btn').addEventListener('click', () => {
            if (!selectedElement) return;
           
            currentCanvas.elements = currentCanvas.elements.filter(el => el.id !== selectedElement.id);
            initCanvas();
            document.getElementById('element-properties-form').style.display = 'none';
            document.getElementById('no-element-selected').style.display = 'block';
        });

        // Anteprima Canvas
        document.getElementById('render-canvas-btn').addEventListener('click', () => {
            const chatbox = document.getElementById('chatbox');
            chatbox.innerHTML += `<div class="assistant">üé® Anteprima Canvas:<br>${JSON.stringify(currentCanvas, null, 2)}</div>`;
            chatbox.scrollTop = chatbox.scrollHeight;
        });

        // Esporta Canvas
        document.getElementById('export-canvas-btn').addEventListener('click', async () => {
            const exportType = document.getElementById('canvas-export-type').value;
           
            try {
                const response = await fetch('/api/export-canvas', {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify({
                        canvas: currentCanvas,
                        type: exportType
                    })
                });
               
                const result = await response.json();
                alert(`Esportazione completata: ${result.url || result.message}`);
            } catch (error) {
                alert(`Errore durante l'esportazione: ${error.message}`);
            }
        });

        // Gestione messaggi esistente (preservata)
        async function handleMessage(userInput) {
            const selectedModel = document.getElementById("api-provider").value;

            if (selectedModel === "ces360") {
                try {
                    const response = await cesAi.chat(userInput);
                    renderMessage("assistant", response.text || "‚ùå Nessuna risposta.");
                } catch (e) {
                    renderMessage("assistant", "‚ùå Errore CES 360 (Puter.js): " + e.message);
                }
            } else {
                try {
                    const res = await fetch("/api/chat", {
                        method: "POST",
                        headers: { "Content-Type": "application/json" },
                        body: JSON.stringify({ message: userInput, api_provider: selectedModel })
                    });
                    const data = await res.json();
                    renderMessage("assistant", data.reply || "‚ùå Nessuna risposta dal backend.");
                } catch (e) {
                    renderMessage("assistant", "‚ùå Errore server: " + e.message);
                }
            }
        }

        function renderMessage(role, text) {
            const chatbox = document.getElementById("chatbox");
            const messageElem = document.createElement("div");
            messageElem.className = role;
            messageElem.innerHTML = `<strong>${role === "user" ? "Tu" : "ArcadiaAI"}:</strong> ${text}`;
            chatbox.appendChild(messageElem);
            chatbox.scrollTop = chatbox.scrollHeight;
        }

        document.getElementById("send-btn").addEventListener("click", () => {
            const input = document.getElementById("input");
            const msg = input.value.trim();
            if (!msg) return;
           
            // Supporto comandi Canvas nella chat
            if (msg.startsWith('@canvas')) {
                const canvasCmd = msg.substring(7).trim();
                handleCanvasCommand(canvasCmd);
                input.value = "";
                return;
            }
           
            renderMessage("user", msg);
            input.value = "";
            handleMessage(msg);
        });

        // Gestione comandi Canvas da chat
        function handleCanvasCommand(command) {
            const chatbox = document.getElementById("chatbox");
           
            if (command === 'nuovo' || command === 'new') {
                currentCanvas = null;
                document.getElementById('canvas-overlay').style.display = 'block';
                initCanvas();
                renderMessage("assistant", "üÜï Nuovo canvas pronto! Usa l'editor per aggiungere contenuti.");
            }
            else if (command.startsWith('aggiungi ') || command.startsWith('add ')) {
                const parts = command.split(' ');
                if (parts.length >= 4) {
                    const type = parts[1];
                    const content = parts.slice(2, -2).join(' ');
                    const x = parseInt(parts[parts.length-2]);
                    const y = parseInt(parts[parts.length-1]);
                   
                    if (!currentCanvas) {
                        currentCanvas = {
                            id: 'canvas-' + Date.now(),
                            elements: []
                        };
                    }
                   
                    currentCanvas.elements.push({
                        id: 'el-' + Date.now(),
                        type: type,
                        content: content,
                        position: { x: x, y: y }
                    });
                   
                    renderMessage("assistant", `‚úÖ Aggiunto elemento ${type} al canvas in (${x},${y})`);
                } else {
                    renderMessage("assistant", "‚ùå Formato comando errato. Usa: @canvas aggiungi [tipo] [contenuto] [x] [y]");
                }
            }
            else {
                renderMessage("assistant", "üîç Scrivi '@canvas nuovo' per iniziare o '@canvas help' per aiuto");
            }
        }
    </script>
</body>
</html>
 """
    return Response(html, content_type="text/html; charset=utf-8")

def chat_with_ces_plus(user_message, conversation_history, attachments=None):
    """
    Versione avanzata di CES con ragionamento passo-passo e messaggi separati.
    """
    if not gemini_model:
        return ["‚ùå CES Plus non √® disponibile al momento."]

    # Prompt identitario forte con istruzioni per il ragionamento
    IDENTITY_PROMPT = (
        "Sei ArcadiaAI CES Plus, la versione avanzata del modello CES con capacit√† di ragionamento approfondito.\n"
        "Prima di rispondere, analizza la domanda e mostra il tuo ragionamento passo-passo.\n"
        "Formatta la risposta in questo modo:\n"
        "[RAGIONAMENTO]\n"
        "(qui il tuo ragionamento logico, passo dopo passo)\n"
        "[RISPOSTA]\n"
        "(qui la risposta finale, ben strutturata)\n"
        "Mantieni un tono professionale ma amichevole."
    )

    try:
        # Costruisci il messaggio completo con allegati
        full_message = user_message if user_message else ""
        if attachments:
            for attachment in attachments:
                if attachment.get('type') == 'application/pdf':
                    try:
                        if isinstance(attachment['data'], str) and attachment['data'].startswith('data:'):
                            file_data = base64.b64decode(attachment['data'].split(',')[1])
                        else:
                            file_data = base64.b64decode(attachment['data'])
                        extracted_text = extract_text_from_file(file_data, 'application/pdf')
                        if extracted_text:
                            full_message += f"\n[CONTENUTO PDF {attachment['name']}]:\n{extracted_text[:10000]}\n"
                    except Exception as e:
                        print(f"Errore elaborazione PDF: {str(e)}")
                        full_message += f"\n[Errore nella lettura del PDF {attachment['name']}]"

        # Prepara la cronologia della conversazione
        contents = []
        contents.append({'role': 'user', 'parts': [{'text': IDENTITY_PROMPT}]})

        # Aggiungi la cronologia recente
        for msg in conversation_history[-4:]:  # Manteniamo meno storia per focalizzarci sul ragionamento
            if isinstance(msg, dict) and 'role' in msg and 'message' in msg:
                role = msg['role'].lower()
                if role == 'user':
                    contents.append({'role': 'user', 'parts': [{'text': msg['message']}]})
                elif role in ['assistant', 'model', 'bot']:
                    contents.append({'role': 'model', 'parts': [{'text': msg['message']}]})

        # Aggiungi il nuovo messaggio con eventuali allegati non PDF
        new_message_parts = [{'text': full_message}] if full_message else []
        if attachments:
            for attachment in attachments:
                mime_type = attachment.get('type', 'application/octet-stream')
                file_name = attachment.get('name', 'file')
                file_data = attachment['data']
                if mime_type == 'application/pdf':
                    continue
                if isinstance(file_data, str) and file_data.startswith('data:'):
                    file_data = file_data.split(',')[1]
                if mime_type.startswith('image/'):
                    new_message_parts.append({
                        'inline_data': {
                            'mime_type': mime_type,
                            'data': file_data,
                            'name': file_name
                        }
                    })
                else:
                    new_message_parts.append({
                        'text': f"[Allegato: {file_name}]"
                    })

        contents.append({'role': 'user', 'parts': new_message_parts})

        # Invia la richiesta a Gemini
        response = gemini_model.generate_content(contents)
        full_reply = response.text

        # Separa il ragionamento dalla risposta finale
        reasoning_part = ""
        answer_part = full_reply
        
        if "[RAGIONAMENTO]" in full_reply and "[RISPOSTA]" in full_reply:
            parts = full_reply.split("[RISPOSTA]")
            reasoning_part = parts[0].replace("[RAGIONAMENTO]", "").strip()
            answer_part = parts[1].strip()
        elif "Ragionamento:" in full_reply:
            parts = full_reply.split("Risposta:", 1)
            reasoning_part = parts[0].replace("Ragionamento:", "").strip()
            answer_part = parts[1].strip() if len(parts) > 1 else full_reply

        # Costruisci la risposta strutturata
        structured_response = []
        if reasoning_part:
            structured_response.append(f"ü§î [Ragionamento]:\n{reasoning_part}")
        structured_response.append(f"üí° [Risposta]:\n{answer_part}")

        return structured_response

    except Exception as e:
        print(f"Errore CES Plus: {str(e)}")
        return ["‚ùå Si √® verificato un errore con CES Plus. Riprova pi√π tardi."]
    except requests.exceptions.Timeout:
        return "‚ùå Timeout del server, riprova pi√π tardi"
    except Exception as e:
        print(f"Errore API Hugging Face: {str(e)}")
        return f"‚ùå Errore temporaneo del servizio: {str(e)}"
    
def format_conversation_history(history):
    """Formatta la cronologia della conversazione per il prompt"""
    if not history:
        return "Nessuna cronologia precedente"
    
    formatted = []
    for msg in history[-6:]:  # Usa solo gli ultimi 6 messaggi per evitare prompt troppo lunghi
        if isinstance(msg, dict):
            role = "Utente" if msg.get("role") == "user" else "Assistente"
            content = msg.get("message", "").strip()
            if content:
                formatted.append(f"{role}: {content}")
    
    return "\n".join(formatted) if formatted else "Nessuna cronologia rilevante"

@app.route("/api/relay-telegram", methods=["POST"])
def relay_telegram():
    # import requests  # <-- RIMUOVI QUESTA RIGA
    data = request.get_json()
    channel_link = data.get("channel_link")
    telegraph_url = data.get("telegraph_url")
    article_text = data.get("article_text")
    match = re.match(r"https://t\.me/([a-zA-Z0-9_]+)", channel_link)
    if not match:
        return jsonify({"success": False, "error": "Link canale non valido"})
    channel_username = match.group(1)
    try:
        from telegram import Bot
        bot = Bot(token=TELEGRAM_TOKEN)
        text = f"üìù Nuovo articolo pubblicato!\n{telegraph_url}\n\nEstratto:\n{article_text[:500]}"
        bot.send_message(chat_id=f"@{channel_username}", text=text)
        return jsonify({"success": True})
    except Exception as e:
        print(f"Errore invio Telegram: {str(e)}")
        return jsonify({"success": False, "error": str(e)})

def meteo_oggi(citt√†):
    """Ottiene le informazioni meteo per una citt√† specifica usando OpenWeatherMap"""
    API_KEY = OPENWEATHERMAP_API_KEY
    if not API_KEY:
        return "‚ùå Errore: API key per OpenWeatherMap non configurata"
    
    try:
        # Codifica la citt√† per l'URL
        citt√†_codificata = urllib.parse.quote(citt√†)
        url = f"http://api.openweathermap.org/data/2.5/weather?q={citt√†_codificata}&appid={API_KEY}&units=metric&lang=it"
        
        response = requests.get(url, timeout=10)
        response.raise_for_status()
        dati = response.json()
        
        if response.status_code != 200 or "weather" not in dati:
            return f"‚ùå Impossibile ottenere il meteo per {citt√†}. La citt√† esiste?"
        
        descrizione = dati["weather"][0]["description"].capitalize()
        temperatura = dati["main"]["temp"]
        umidit√† = dati["main"]["humidity"]
        vento = dati["wind"]["speed"]
        
        return (
            f"‚õÖ Meteo a {citt√†}:\n"
            f"- Condizioni: {descrizione}\n"
            f"- Temperatura: {temperatura}¬∞C\n"
            f"- Umidit√†: {umidit√†}%\n"
            f"- Vento: {vento} m/s"
        )
    
    except requests.exceptions.RequestException as e:
        print(f"Errore API meteo: {str(e)}")
        return f"‚ùå Errore temporaneo nel servizio meteo. Riprova pi√π tardi."
    except Exception as e:
        print(f"Errore generico meteo: {str(e)}")
        return f"‚ùå Si √® verificato un errore nel recupero del meteo per {citt√†}."
    
    
def parse_quick_command(input_text):
    """
    Analizza un comando rapido con prefisso @ e restituisce una tupla (comando, argomento).
    Se non √® un comando rapido, restituisce (None, None).
    
    Esempio:
    >>> parse_quick_command("@cerca seconda guerra mondiale")
    ('cerca', 'seconda guerra mondiale')
    >>> parse_quick_command("ciao come stai?")
    (None, None)
    """
    if not input_text.startswith("@"):
        return None, None
    
    # Rimuovi il @ iniziale e dividi il resto in comando e argomento
    parts = input_text[1:].split(" ", 1)  # Dividi al primo spazio
    command = parts[0].lower().strip()
    argument = parts[1].strip() if len(parts) > 1 else ""
    
    return command, argument

def search_flathub_apps(query):
    """Cerca app su Flathub e restituisce i risultati"""
    url = f"https://flathub.org/api/v1/apps/search/{query}"
    try:
        response = requests.get(url, timeout=10)
        response.raise_for_status()
        return response.json()
    except Exception as e:
        print(f"Errore ricerca Flathub: {str(e)}")
        return None

def get_flathub_app_details(app_id):
    """Ottiene i dettagli di un'app da Flathub"""
    url = f"https://flathub.org/api/v1/apps/{app_id}"
    try:
        response = requests.get(url, timeout=10)
        response.raise_for_status()
        return response.json()
    except Exception as e:
        print(f"Errore dettagli app Flathub: {str(e)}")
        return None

def get_flathub_download_link(app_id):
    """Genera il link di download diretto per un'app Flathub"""
    return f"https://dl.flathub.org/repo/appstream/{app_id}.flatpakref"

def handle_flathub_command(command, argument):
    """Gestisce i comandi relativi a Flathub"""
    if command == "flathub":
        if not argument:
            return "‚ùå Specifica un'app da cercare su Flathub. Esempio: @flathub firefox"
        
        # Cerca l'app su Flathub
        results = search_flathub_apps(argument)
        if not results:
            return f"‚ùå Nessun risultato trovato per '{argument}' su Flathub"
        
        # Prendi il primo risultato
        app = results[0]
        app_id = app.get("flatpakAppId")
        app_name = app.get("name", "Sconosciuto")
        
        # Ottieni il link di download
        download_link = get_flathub_download_link(app_id)
        
        return (
            f"üì¶ {app_name} su Flathub:\n"
            f"üîó Download diretto: {download_link}\n"
            f"‚ÑπÔ∏è Dettagli: https://flathub.org/apps/{app_id}"
        )
    
    elif command == "flathub_download":
        if not argument:
            return "‚ùå Specifica l'ID dell'app da scaricare. Esempio: @flathub_download org.mozilla.firefox"
        
        # Verifica se l'app esiste
        app_details = get_flathub_app_details(argument)
        if not app_details:
            return f"‚ùå App con ID '{argument}' non trovata su Flathub"
        
        # Genera il link di download
        download_link = get_flathub_download_link(argument)
        
        return (
            f"‚¨áÔ∏è Download diretto per {app_details.get('name', argument)}:\n"
            f"{download_link}\n\n"
            f"Per installare: flatpak install --from {download_link}"
        )
    
    return None
# --- Verifica disponibilit√† di Winget all'avvio (globale) ---
_WINGET_AVAILABLE = False
try:
    subprocess.run(['winget', '--version'], capture_output=True, text=True, check=True)
    _WINGET_AVAILABLE = True
    print("Winget √® disponibile sul sistema.")
except (subprocess.CalledProcessError, FileNotFoundError):
    print("ATTENZIONE: Winget non √® disponibile. Le funzioni Winget potrebbero non funzionare.")

# --- Verifica disponibilit√† di Winget all'avvio (globale) ---
def search_windows_apps_online(query):
    """
    Cerca app Windows online usando una ricerca web per trovare link di download ufficiali.
    """
    search_term = f"{query} download ufficiale Windows"
    results = search_duckduckgo(search_term) # Usa la tua funzione esistente per la ricerca web

    if not results:
        return []

    # Filtra e prioritizza i risultati per trovare link di download ufficiali.
    # Questa √® una euristica e pu√≤ essere migliorata per la tua specifica necessit√†.
    filtered_urls = []
    official_keywords = ["microsoft.com", "google.com", "mozilla.org", "videolan.org", "spotify.com", "7-zip.org"]
    winget_web_interfaces = ["winget.run", "winstall.app"] # Siti che mostrano info Winget

    for url in results:
        # Priorit√† 1: Siti ufficiali di software molto noti o Winget web interfaces
        if any(keyword in url.lower() for keyword in official_keywords + winget_web_interfaces):
            filtered_urls.append(url)
        # Priorit√† 2: Siti che includono "download" e "official" nel percorso, evitando aggregatori noti
        elif "download" in url.lower() and not any(bad_domain in url for bad_domain in ["softonic.com", "filehippo.com", "updatestar.com"]):
            filtered_urls.append(url)
    
    # Rimuovi duplicati mantenendo l'ordine originale
    seen = set()
    unique_filtered_urls = []
    for url in filtered_urls:
        if url not in seen:
            unique_filtered_urls.append(url)
            seen.add(url)

    return unique_filtered_urls[:3] # Restituisce i primi 3 URL pi√π rilevanti

def get_windows_app_details_from_url(url, original_query):
    """
    Tenta di estrarre dettagli base da una URL di download per un'app Windows.
    Questo √® un approccio semplificato dato che non abbiamo i dati strutturati di Winget.
    """
    name = original_query.capitalize() # Nome predefinito capitalizzato dalla query
    is_proprietary = True # Predefinito a proprietario, poi proviamo a indovinare
    publisher = "Sconosciuto (verificare pagina)"

    # Heuristiche basate su URL per nome e licenza
    if "mozilla.org/firefox" in url.lower():
        name = "Mozilla Firefox"
        is_proprietary = False # Firefox √® open source
        publisher = "Mozilla"
    elif "google.com/chrome" in url.lower():
        name = "Google Chrome"
        is_proprietary = True
        publisher = "Google LLC"
    elif "microsoft.com/edge" in url.lower():
        name = "Microsoft Edge"
        is_proprietary = True
        publisher = "Microsoft Corporation"
    elif "videolan.org/vlc" in url.lower():
        name = "VLC Media Player"
        is_proprietary = False
        publisher = "VideoLAN"
    elif "7-zip.org" in url.lower():
        name = "7-Zip"
        is_proprietary = False
        publisher = "Igor Pavlov"
    elif "spotify.com" in url.lower():
        name = "Spotify"
        is_proprietary = True
        publisher = "Spotify AB"
    elif "winget.run" in url.lower() or "winstall.app" in url.lower():
        # Se √® un'interfaccia Winget, proviamo a estrarre il nome dalla query
        name = original_query.capitalize() + " (dal catalogo Winget)"
        publisher = "Vari (dipende dall'app)"


    return {
        'name': name,
        'homepage': url,          # L'URL di riferimento √® la homepage/pagina di download
        'installer_url': url,     # In questo contesto, √® la stessa cosa
        'is_proprietary': is_proprietary,
        'publisher': publisher
    }

def handle_winget_command(command, argument):
    """
    Gestisce i comandi relativi alle app Windows tramite ricerca web.
    """
    if command == "winget":
        if not argument:
            return "‚ùå Specifica un'app Windows da cercare. Esempio: @winget firefox"
        
        # Ora cerchiamo online invece di eseguire winget localmente
        download_urls = search_windows_apps_online(argument)

        if not download_urls:
            return (
                f"‚ùå Non sono riuscito a trovare un link di download ufficiale per '{argument}' per Windows.\n"
                "Potresti provare a cercare manualmente su Google o DuckDuckGo."
            )
        
        # Prendo il primo risultato come il pi√π rilevante
        main_url = download_urls[0]
        app_details = get_windows_app_details_from_url(main_url, argument) # Usa una funzione semplificata per estrarre i dettagli

        name = app_details['name']
        homepage = app_details['homepage']
        is_proprietary = app_details['is_proprietary']
        publisher = app_details['publisher']
        
        license_note = ""
        if is_proprietary:
            license_note = (
                f"Nota: **{name}** √® un software proprietario. "
                f"La licenza d'uso √® definita dall'editore ({publisher})."
            )
        else:
            license_note = (
                f"Nota: **{name}** √® un software open source/gratuito. "
                f"La licenza d'uso √® definita dall'editore ({publisher})."
            )

        response = f"""
üì¶ **{name}** per Windows:

1.  **Scarica l'installer:**
    Visita la pagina ufficiale per il download: {homepage}
    *(Dovrai scaricare e installare l'app manualmente dal sito.)*

‚ÑπÔ∏è Pagina ufficiale dell'app: {homepage}

{license_note}
        """
        return response.strip()
    
    return None # Comando Winget non riconosciuto (es. se avessi altri sottocomandi)
# --- Nuove Funzioni per Snap Store ---
def search_snap_store_online(query):
    """
    Cerca app Snap online usando una ricerca web per trovare link ufficiali su snapcraft.io.
    """
    app.logger.info(f"Cercando app Snap online per: {query}")
    search_term = f"{query} snapcraft.io"
    results = search_duckduckgo(search_term)
    return [url for url in results if "snapcraft.io" in url.lower()][:3]

def get_snap_app_details_from_url(url, original_query):
    """
    Estrae dettagli base da una URL di Snapcraft per un'app Snap.
    """
    name = original_query.capitalize() + " (Snap)"
    is_proprietary = "proprietary" in url.lower() # Semplice euristica per licenza
    publisher = "Snapcraft Community / Canonical" # Default

    # Migliora l'estrazione del nome e del publisher se possibile
    if "snapcraft.io/store" not in url.lower():
        try:
            # Estrai il nome dal path dell'URL (es. snapcraft.io/firefox -> firefox)
            name_from_url = url.split('/')[-1].replace('-', ' ').title()
            if name_from_url and len(name_from_url) > 1: # Assicura che non sia vuoto o solo un carattere
                name = name_from_url + " (Snap)"
        except Exception as e:
            app.logger.warning(f"Impossibile estrarre nome Snap da URL {url}: {e}")
            
    # Euristiche per alcune app comuni
    if original_query.lower() == "spotify":
        name = "Spotify (Snap)"
        is_proprietary = True
        publisher = "Spotify"
    elif original_query.lower() == "vlc":
        name = "VLC media player (Snap)"
        is_proprietary = False
        publisher = "VideoLAN"

    return {
        'name': name,
        'homepage': url,
        'installer_info': f"Installazione via Snap: `sudo snap install {original_query.lower()}`", # Comando di installazione Snap
        'is_proprietary': is_proprietary,
        'publisher': publisher,
        'platform': 'Linux (Snap)'
    }

def handle_snap_command(argument):
    """
    Gestisce i comandi relativi alle app Snap.
    """
    if not argument:
        return "‚ùå Specifica un'app Snap da cercare. Esempio: @snap spotify"
    
    download_urls = search_snap_store_online(argument)

    if not download_urls:
        return (
            f"‚ùå Non sono riuscito a trovare un pacchetto Snap ufficiale per '{argument}'.\n"
            "Potrebbe non esistere o il nome non √® corretto. Prova a cercare su snapcraft.io."
        )
    
    main_url = download_urls[0]
    app_details = get_snap_app_details_from_url(main_url, argument)

    name = app_details['name']
    homepage = app_details['homepage']
    installer_info = app_details['installer_info']
    is_proprietary = app_details['is_proprietary']
    publisher = app_details['publisher']
    platform = app_details['platform']
    
    license_note = ""
    if is_proprietary:
        license_note = (
            f"Nota: **{name}** √® un software proprietario. "
            f"La licenza d'uso √® definita dall'editore ({publisher})."
        )
    else:
        license_note = (
            f"Nota: **{name}** √® un software open source/gratuito. "
            f"La licenza d'uso √® definita dall'editore ({publisher})."
        )

    response = f"""
üì¶ **{name}** per {platform}:

1.  **Pagina ufficiale:** {homepage}
2.  **Installazione:** `{installer_info}`
    *(Apri il terminale sul tuo sistema Linux e incolla il comando.)*

{license_note}
    """
    return response.strip()


def search_fdroid_online(query):
    """
    Cerca app F-Droid online usando una ricerca web per trovare link ufficiali su f-droid.org.
    """
    app.logger.info(f"Cercando app F-Droid online per: {query}")
    search_term = f"{query} f-droid.org"
    results = search_duckduckgo(search_term)
    # Filtra per URL che sembrano pagine di pacchetto F-Droid (es. /packages/nome.app/)
    return [url for url in results if "f-droid.org/packages/" in url.lower()][:3]

def get_fdroid_app_details_from_url(url, original_query):
    """
    Estrae dettagli base da una URL di F-Droid per un'app Android.
    """
    name = original_query.capitalize() + " (F-Droid)"
    is_proprietary = False # Le app su F-Droid sono sempre open source
    publisher = "F-Droid Community / Sviluppatore originale" # Default

    try:
        # Estrai il nome del pacchetto dall'URL per un nome pi√π preciso
        match = re.search(r"f-droid\.org/packages/([a-zA-Z0-9\._-]+)/", url)
        if match:
            package_id = match.group(1)
            # Tenta di pulire il nome del pacchetto per renderlo pi√π leggibile
            cleaned_name = package_id.split('.')[-1].replace('_', ' ').title()
            if cleaned_name:
                name = cleaned_name + " (F-Droid)"
    except Exception as e:
        app.logger.warning(f"Impossibile estrarre nome F-Droid da URL {url}: {e}")

    # Euristiche per alcune app comuni
    if original_query.lower() == "newpipe":
        name = "NewPipe (F-Droid)"
        publisher = "Team NewPipe"
    elif original_query.lower() == "signal":
        name = "Signal (F-Droid)"
        publisher = "Signal Messenger, LLC"

    return {
        'name': name,
        'homepage': url,
        'installer_info': f"Scarica direttamente da F-Droid.org o tramite l'app F-Droid.",
        'is_proprietary': is_proprietary, # Sempre False per F-Droid
        'publisher': publisher,
        'platform': 'Android (F-Droid)'
    }

def handle_fdroid_command(argument):
    """
    Gestisce i comandi relativi alle app F-Droid.
    """
    try: # Aggiunto try-except
        if not argument:
            return "‚ùå Specifica un'app F-Droid da cercare. Esempio: @fdroid newpipe"
        
        download_urls = search_fdroid_online(argument)

        if not download_urls:
            return (
                f"‚ùå Non sono riuscito a trovare un'app ufficiale per '{argument}' su F-Droid.\n"
                "Potrebbe non esistere o il nome non √® corretto. Prova a cercare su f-droid.org."
            )
        
        main_url = download_urls[0]
        app_details = get_fdroid_app_details_from_url(main_url, argument)

        name = app_details['name']
        homepage = app_details['homepage']
        installer_info = app_details['installer_info']
        is_proprietary = app_details['is_proprietary']
        publisher = app_details['publisher']
        platform = app_details['platform']
        
        license_note = (
            f"Nota: **{name}** √® un software open source/gratuito. "
            f"La licenza d'uso √® definita dall'editore ({publisher})."
        )

        response = f"""
üì¶ **{name}** per {platform}:

1.  **Pagina ufficiale:** {homepage}
2.  **Installazione:** {installer_info}
    *(Dovrai scaricare e installare l'app manualmente sul tuo dispositivo Android o usare l'app F-Droid.)*

{license_note}
        """
        return response.strip()
    except Exception as e:
        app.logger.error(f"Errore nella gestione del comando F-Droid per '{argument}': {str(e)}")
        return f"‚ùå Si √® verificato un errore interno durante l'elaborazione del comando F-Droid. Controlla i log del server per maggiori dettagli."

# --- Funzione principale della tua IA che elabora l'input dell'utente ---
def process_user_input(user_input):
    """
    Funzione principale della tua IA che elabora l'input dell'utente.
    """
    if user_input.lower() == "@app":
        return """
Repository per Download Manager:
- @Flathub - [Nome-App] (per Linux)
- @Winget - [Nome-App] (per Windows)
- @Snap - [Nome-App] (per Linux)
- @F-Droid - [Nome-App] (per Android)
(Nota: Su Winget sono disponibili sia app gratuite che a pagamento, e con licenze diverse. Il Download Manager fornisce l'installer, ma la licenza d'uso √® definita dall'editore dell'app.)
"""
 
    elif user_input.lower().startswith("@winget"): # Usiamo direttamente user_input.lower()
        parts = user_input.lower().split(" ", 1)
        argument = parts[1].strip() if len(parts) > 1 else ""
        return handle_winget_command("winget", argument)
    
    elif user_input.lower().startswith("@snap"): # Usiamo direttamente user_input.lower()
        parts = user_input.lower().split(" ", 1)
        argument = parts[1].strip() if len(parts) > 1 else ""
        return handle_snap_command(argument)
    
    elif user_input.lower().startswith("@fdroid"): # Usiamo direttamente user_input.lower()
        parts = user_input.lower().split(" ", 1)
        argument = parts[1].strip() if len(parts) > 1 else ""
        return handle_fdroid_command(argument)
    
    elif user_input.lower().startswith("@flathub"): # Usiamo direttamente user_input.lower()
        parts = user_input.lower().split(" ", 1)
        argument = parts[1].strip() if len(parts) > 1 else ""
        return handle_flathub_command(argument)
    
    else:
        return "Comando non riconosciuto. Prova '@app' per vedere i repository disponibili."

from telegram import Bot
from telegram.error import TelegramError


@app.route('/invia-telegram', methods=['POST'])
def invia_telegram():
    data = request.get_json()
    chat_id = data.get('chat_id')  # Es: "@Leoniaplusgiornale"
    text = data.get('text')        # Cambiato da 'testo' a 'text'
    link_telegraph = data.get('link_telegraph')

    if not all([chat_id, text, link_telegraph]):
        return jsonify({"success": False, "error": "Dati mancanti"}), 400

    try:
        bot = Bot(token=TELEGRAM_TOKEN)
        bot.send_message(
            chat_id=chat_id,
            text=f"üì¢ Nuovo articolo!\n\n{text[:1000]}...\n\nLeggi tutto: {link_telegraph}",
            parse_mode='HTML',
            disable_web_page_preview=False
        )
        return jsonify({"success": True})
    except Exception as e:
        print(f"ERRORE TELEGRAM: {str(e)}")
        return jsonify({"success": False, "error": str(e)})
    
from flask import session
import requests
import locale
import json
import re
import datetime
import base64
def handle_quick_commands(message, experimental_mode=False, attachments=None, api_provider="gemini", conversation_history=None):
    """
    Gestisce i comandi rapidi di ArcadiaAI.
    Nota: Questa funzione √® ora asincrona (`async def`) a causa della chiamata a `await invia_a_telegram`.
    """
    msg_lower = message.strip().lower()
    command, argument = parse_quick_command(message)
    if command is None:
        return None

    print(f"DEBUG: Comando ricevuto - command: {command}, argument: {argument}")  # Debug logging


    # Prima verifica i comandi speciali
    if msg_lower == "@impostazioni modalit√† sperimentale disattiva":
        return "‚ùé Modalit√† sperimentale disattivata!"

    if msg_lower == "@impostazioni modalit√† sperimentale attiva":
        return "‚úÖ Modalit√† sperimentale attivata!"

    # Poi gestisci i comandi principali
    if command == "cerca":
        if not argument:
            return "‚ùå Devi specificare cosa cercare. Esempio: @cerca seconda guerra mondiale"

        try:
            results = search_web(argument)
            if not results:
                return f"‚ùå Nessun risultato trovato per '{argument}'"

            response = f"üîç Risultati per '{argument}':\n\n"
            for i, result in enumerate(results[:3], 1):
                response += f"{i}. {result.get('title', 'Nessun titolo')}\n"
                response += f"   {result.get('url', 'Nessun URL')}\n"
                response += f"   {result.get('snippet', 'Nessuna descrizione')}\n\n"

            return response.strip()

        except Exception as e:
            print(f"Errore durante la ricerca: {str(e)}")
            return "‚ùå Errore temporaneo durante la ricerca. Riprova pi√π tardi."

    elif command == "telegraph":
        if not argument:
            return "‚ùå Devi specificare il contenuto da pubblicare. Esempio: @telegraph Questo √® un articolo di prova"

        # Se l'argomento inizia con "scrivi" o "scrivimi", genera contenuto con AI
        if argument.lower().startswith(("scrivi", "scrivimi")):
            try:
                # Genera il contenuto con Gemini
                generated_content = chat_with_gemini(argument, conversation_history or [])
                if not generated_content:
                    return "‚ùå Impossibile generare il contenuto"

                # Pubblica su Telegraph
                title = "Articolo generato da ArcadiaAI"
                telegraph_url = publish_to_telegraph(title, generated_content)

                if telegraph_url and not telegraph_url.startswith("‚ö†Ô∏è"):
                    return f"üìù Articolo pubblicato su Telegraph: {telegraph_url}"
                return telegraph_url or "‚ùå Errore nella pubblicazione"

            except Exception as e:
                print(f"Errore generazione/pubblicazione: {str(e)}")
                return "‚ùå Errore durante la generazione/pubblicazione"
        else:
            title = "Articolo generato da ArcadiaAI"
            telegraph_url = publish_to_telegraph(title, argument)
            return f"üìù Pubblicato su Telegraph: {telegraph_url}" if telegraph_url else "‚ùå Errore pubblicazione"

    elif command == "meteo" and argument:
        return meteo_oggi(argument)

    elif command == "data":
        locale.setlocale(locale.LC_TIME, "it_IT.UTF-8")
        oggi = datetime.datetime.now().strftime("%A %d %B %Y")
        return f"üìÖ Oggi √® {oggi}"

    elif command == "source":
        return "üîó Repository GitHub: https://github.com/Mirko-linux/Nova-App/tree/main/ArcadiaAI"

    elif command == "arcadia":
        return "üîó Profilo della Repubblica di Arcadia: https://t.me/Repubblica_Arcadia"

    elif msg_lower == "@impostazioni lingua italiano":
        return "üáÆüáπ Lingua cambiata in italiano!"

    elif msg_lower == "@impostazioni lingua inglese":
        return "üá¨üáß Lingua cambiata in inglese!"

    elif command == "app":
        return (
            "Repository per download manager:\n"
            "- @flathub - [nome-app]: scarica un'app per Linux(flatpak)\n"
            "- @winget - [nome-app]: scarica un'app per Windows\n"
            "- @snap - [nome-app]: scarica un'app per Linux (Snapd)\n"
            "- @fdroid - [nome-app]: scarica un'app per Android (F-Droid)\n"
            "(Nota: Su Winget sono disponibili sia app gratuite che a pagamento, "
            "e con licenze diverse. Il Download Manager fornisce l'installer, "
            "ma la licenza d'uso √® definita dall'editore dell'app.)"
        )

    elif command == "flathub":
        if not argument:
            return "‚ùå Specifica un'app da cercare su Flathub. Esempio: @flathub firefox"
        return handle_flathub_command("flathub", argument)
    elif command == "snap":
        if not argument:
            return "‚ùå Specifica un'app Snap da cercare. Esempio: @snap firefox"
        return handle_snap_command(argument)

    elif command == "fdroid":
        if not argument:
            return "‚ùå Specifica un'app F-Droid da cercare. Esempio: @fdroid newpipe"
        return handle_fdroid_command(argument)

    elif command == "tos":
        return "üìú Termini di Servizio: https://arcadiaai.netlify.app/documentazioni"

    elif command == "codice_sorgente":
        return "üîó Codice sorgente di ArcadiaAI: https://github.com/Mirko-linux/Nova-App/tree/main/ArcadiaAI"

    elif command == "info":
        return (
            "‚ÑπÔ∏è Informazioni su ArcadiaAI:\n\n"
            "Versione: 1.5.7\n"
            "Modello di testo: CES 1.5 v2\n"
            "Modello di Ragionamento: CES Plus\n"
            "Modello esterni: DeepSeek Chat\n"
            "Lingua: Italiano e inglese (beta)\n"
            "Creatore: Mirko Yuri Donato\n"
            "Licenza: Mozilla Public License 2.0\n"
            "Repository: https://github.com/Mirko-linux/Nova-App/tree/main/ArcadiaAI\n"
            "Termini di Servizio: https://arcadiaai.netlify.app/documentazioni"
        )

    elif command == "crea" and argument and argument.lower().startswith("zip"):
        return "Per creare uno ZIP allega i file e usa il comando dalla chat. Il file ZIP verr√† generato dal frontend."

    elif command == "impostazioni":
        return (
            "‚öôÔ∏è Menu Impostazioni:\n\n"
            "- Modalit√† Sperimentale: attiva/disattiva\n"
            "- Lingua: italiano/inglese\n"
            "- Tema: chiaro/scuro\n"
            "- Modalit√† Sviluppatore: attiva/disattiva\n"
            "Usa i comandi @impostazioni [opzione] per modificare le impostazioni.\n"
            "Nota: Alcune opzioni potrebbero non essere disponibili in questa versione."
        )

    elif command == "immagine":
        if not argument:
            return "‚ùå Specifica cosa disegnare. Esempio: @immagine drago pixel art su un hoverboard"
        try:
            response = requests.post(
                "https://arcadiaai.onrender.com/api/ces-image",
                headers={"Content-Type": "application/json"},
                json={"prompt": argument},
                timeout=20
            )
            if response.status_code == 200:
                data = response.json()
                image_url = data.get("image_url")
                if image_url:
                    return (
                        f'<div style="text-align: center;">'
                        f'<img src="{image_url}" alt="{argument}" '
                        f'style="max-width: 512px; width: 100%; height: auto; '
                        f'margin-top: 8px; border-radius: 6px;">'
                        f'</div>'
                    )
                return "‚ö†Ô∏è L'API ha risposto ma non ha fornito nessuna immagine."
            return f"‚ùå Errore dall'API CES Image: {response.status_code}"
        except Exception as e:
            return f"‚ùå Errore nella generazione dell'immagine: {e}"

    elif command == "musica":
        if not argument:
            return "‚ùå Specifica che tipo di musica generare. Esempio: @musica synthwave futuristica"

        # LISTA DEGLI ENDPOINT PUBBLICI E GRATUITI
        endpoints = [
            "https://suno-api-2025.free.hf.space/api/generate", # Spesso offline o sovraccarico
            "https://arcadiaai.onrender.com/api/generate-musica", # Il tuo endpoint personale
            "https://musicgen-proxy-2025.fly.dev/generate" # Probabilmente non pi√π attivo
        ]

        for endpoint in endpoints:
            print(f"DEBUG: Tento di usare l'endpoint -> {endpoint}")
            try:
                response = requests.post(
                    endpoint,
                    headers={"Content-Type": "application/json"},
                    json={"prompt": argument},
                    timeout=30 # Timeout leggermente pi√π lungo per i servizi lenti
                )

                if response.status_code == 200:
                    print(f"DEBUG: Risposta 200 OK da {endpoint}")
                    try:
                        data = response.json()
                        music_url = data.get("music_url") or data.get("url") or data.get("audio_url")

                        if music_url and isinstance(music_url, str) and music_url.startswith('http'):
                            print(f"DEBUG: Successo! URL trovato: {music_url}")

                            # --- NUOVA LOGICA PER IL TIPO DI AUDIO ---
                            audio_type = "audio/mpeg" # Default per MP3
                            if music_url.lower().endswith('.wav'):
                                audio_type = "audio/wav"
                            elif music_url.lower().endswith('.ogg'):
                                audio_type = "audio/ogg"
                            # Puoi aggiungere altri tipi se necessario (es. .aac, .flac)

                            return (
                                f'<div style="text-align: center; margin: 15px 0;">'
                                f'<audio controls style="width: 100%; max-width: 400px; border-radius: 12px; box-shadow: 0 4px 8px rgba(0,0,0,0.1);">'
                                f'<source src="{music_url}" type="{audio_type}">' # Usa il tipo di audio dinamico
                                f'Il tuo browser non supporta l\'elemento audio.'
                                f'</audio><br>'
                                f'<small style="color: #666; font-size: 0.8em;">Generato tramite endpoint pubblico</small>'
                                f'</div>'
                            )
                        else:
                            print(f"DEBUG: Risposta OK ma URL non valido. Provo il prossimo.")
                            continue # Prova il prossimo servizio
                    except ValueError:
                        print(f"DEBUG: Risposta non √® un JSON valido. Provo il prossimo.")
                        continue # Prova il prossimo servizio
                else:
                    print(f"DEBUG: Errore da {endpoint} (Status: {response.status_code}). Provo il prossimo.")
                    continue # Prova il prossimo servizio

            except requests.exceptions.RequestException as e:
                print(f"DEBUG: Endpoint {endpoint} non raggiungibile ({e}). Provo il prossimo.")
                continue # Prova il prossimo servizio

        # --- FALLBACK ---
        print("DEBUG: Nessun endpoint ha funzionato. Attivo il fallback manuale.")
        return (
            '‚ö†Ô∏è <b>Nessun servizio di generazione automatica √® al momento disponibile.</b><br><br>'
            'Puoi comunque creare la tua musica manualmente tramite il link qui sotto:'
            '<div style="text-align: center; margin-top: 15px;">'
            f'<a href="https://huggingface.co/spaces/facebook/MusicGen" target="_blank" style="display: inline-block; padding: 12px 20px; background-color: #007bff; color: white; text-decoration: none; border-radius: 8px; font-weight: 500;">'
            'üéß Genera Musica su MusicGen'
            '</a>'
            '</div>'
        )
    elif command == "privacy":
        return (
            "üîí Privacy Policy:\n\n"
            "I tuoi dati non vengono memorizzati o condivisi. "
            "Le conversazioni sono salvate in locale. "
            "Per maggiori dettagli, consulta i Termini di Servizio."
        )

    elif command == "sac":
        return (
            "üîó SAC (Strumenti Avanzati di CES):\n\n"
            "I SAC sono componenti interne del codice sorgente di ArcadiaAI"
        )

    elif command == "telegram":
        # Estrai il link del canale e il prompt
        parts = argument.split("prompt", 1)
        if len(parts) < 2:
            return "‚ùå Formato errato. Usa: @telegram [link_canale] prompt \"testo\""

        channel_link = parts[0].strip()
        prompt_text = parts[1].strip().strip('"')

        # Verifica il formato del link
        if not re.match(r"https://t\.me/[a-zA-Z0-9_]+", channel_link):
            return "‚ùå Link canale non valido. Esempio: https://t.me/nomecanale"

        # Genera articolo con Gemini
        try:
            article, telegraph_url = generate_with_gemini(prompt_text, "Articolo da ArcadiaAI")
            if not telegraph_url or telegraph_url.startswith("‚ö†Ô∏è"):
                return f"‚ùå Errore nella pubblicazione su Telegraph: {telegraph_url}"

            # Estrai l'username del canale dal link
            channel_username = channel_link.split('/')[-1]

            # Invia al canale Telegram
            response = requests.post(
                "https://arcadiaai.onrender.com/invia-telegram",  # Assicurati che questo sia il tuo URL corretto
                json={
                    "chat_id": f"@{channel_username}",
                    "text": f"<b>üì¢ Nuovo articolo pubblicato!</b>\n\n{article[:500]}...",  # Cambiato da 'testo' a 'text'
                    "link_telegraph": telegraph_url
                },
                timeout=20
            )

            if response.status_code == 200:
                result = response.json()
                if result.get("success"):
                    return f"‚úÖ Articolo pubblicato e inviato correttamente!\n{telegraph_url}"
                else:
                    return f"‚ö†Ô∏è Articolo pubblicato su Telegraph ma errore invio Telegram: {result.get('error', 'Errore sconosciuto')}\n{telegraph_url}"
            else:
                return f"‚ö†Ô∏è Errore nella richiesta al server: {response.status_code}"

        except Exception as e:
            return f"‚ùå Errore durante l'operazione: {str(e)}"

    elif command == "riassumi":
        # Assumi che extract_text_from_file, chat_with_gemini, chat_with_ces_plus,
        # e chat_with_deepseek siano definiti altrove
        # E gemini_model sia accessibile in questo scope se api_provider √® "gemini" o "cesplus"
        if attachments:
            for attachment in attachments:
                if attachment.get('type') == 'application/pdf':
                    file_data = base64.b64decode(
                        attachment['data'].split(',')[1] if attachment['data'].startswith('data:') else attachment['data']
                    )
                    testo = extract_text_from_file(file_data, 'application/pdf')
                    if not testo:
                        return "‚ùå Impossibile estrarre testo dal PDF."
                    prompt = f"Riassumi il seguente testo:\n\n{testo[:12000]}"
                    if api_provider == "gemini" and 'gemini_model' in locals():
                        return chat_with_gemini(prompt, conversation_history)
                    elif api_provider == "cesplus" and 'gemini_model' in locals():
                        return chat_with_ces_plus(prompt, conversation_history)
                    elif api_provider == "deepseek":
                        return chat_with_deepseek(prompt, conversation_history)
                    else:
                        return "‚ùå Provider non riconosciuto o modello non disponibile."
            return "‚ùå Nessun PDF trovato tra gli allegati."
        if not argument or not argument.startswith("http"):
            return "‚ùå Allega un file PDF o fornisci un URL diretto al PDF da riassumere."

    elif command == "aiuto":
        return (
            "üéØ Comandi rapidi disponibili:\n\n"
            "@cerca [query] - Cerca su internet\n"
            "@telegraph [testo|url] - Pubblica testo o il contenuto di un URL su Telegraph\n"
            "@meteo [luogo] - Ottieni il meteo\n"
            "@data - Mostra la data di oggi\n"
            "@aiuto - Mostra questa guida\n"
            "@impostazioni modalit√† sperimentale disattiva - Disattiva la modalit√† sperimentale\n"
            "@impostazioni - apre il menu delle impostazioni\n"
            "@impostazioni lingua [nome lingua] - Cambia lingua\n"
            "@impostazioni tema [chiaro|scuro] - Cambia tema\n"
            "@impostazioni modalit√† sviluppatore attiva - Attiva la modalit√† sviluppatore\n"
            "@ToS - Mostra i Termini di Servizio\n"
            "@Arcadia - Mostra il profilo della Repubblica di Arcadia\n"
            "@info - Mostra informazioni su ArcadiaAI\n"
            "@impostazioni modalit√† sperimentale attiva - Attiva la modalit√† sperimentale\n"
            "@codice_sorgente - Mostra il codice sorgente di ArcadiaAI\n"
            "@impostazioni - Mostra il menu delle impostazioni\n"
            "@estensioni - Mostra le estensioni installate\n"
            "@estensioni [nome estensione] - Mostra informazioni su un'estensione\n"
            "@privacy - Mostra la privacy policy\n"
            "@app - Mostra le repository che supportano il Download Manager\n"
            "@flathub [nome app] - Cerca un'app su Flathub e restituisce il download diretto\n"
            "@flathub_download [ID app] - Scarica un'app specifica da Flathub\n"
            "@winget [nome app] - Cerca un'app su Winget e restituisce il download diretto (NOVIT√Ä!)\n"
            "@cesplus - Usa il modello CES Plus per risposte avanzate\n"
            "@crea zip - genera un file ZIP con file dati dall'utente\n"
            "@esporta - Esporta l'ultima conversazione in un file TXT\n"
            "@importa - Importa la una conversazione da un file TXT tramite NovaSync\n"
            "@riassumi https://www.ilovepdf.com/ - Legge e riassume un file PDF\n"
            "Per altre domande, chiedi pure!"
        )

    # Se nessun comando √® riconosciuto
    return f"‚ùå Comando '{command}' non riconosciuto. Scrivi '@aiuto' per la lista dei comandi."

def should_use_predefined_response(message):
    """Determina se usare una risposta predefinita solo per domande molto specifiche"""

    message = message.lower().strip()
    exact_matches = [
        "chi sei esattamente",
        "qual √® la tua identit√† precisa",
        "elencami tutte le tue funzioni",
        # ... altre domande esatte
    ]
    return any(exact_match in message for exact_match in exact_matches)
nest_asyncio.apply()

# === Memoria contestuale per utente ===
user_histories = {}  # user_id: [messaggi]
user_models = {}     # user_id: modello selezionato

# === HANDLER MESSAGGI TELEGRAM ===
async def telegram_message_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    message = update.message.text
    chat_id = update.message.chat.id
    user_id = update.effective_user.id

    # Recupera o inizializza la cronologia
    history = user_histories.get(user_id, [])
    history.append({"role": "user", "message": message})

    # Recupera il modello selezionato o usa "gemini" come default
    modello = user_models.get(user_id, "gemini")

    data = {
        "message": message,
        "conversation_history": history,
        "api_provider": modello
    }

    try:
        response = requests.post("https://arcadiaai.onrender.com/chat", json=data, timeout=60)
        reply = response.json().get("reply", "‚ùå Nessuna risposta generata.")
    except Exception as e:
        reply = f"‚ùå Errore: {e}"

    # Salva la risposta nella cronologia
    history.append({"role": "assistant", "message": reply})
    user_histories[user_id] = history[-2000:]  # Mantieni solo gli ultimi 20 messaggi

    await context.bot.send_message(chat_id=chat_id, text=reply)

# === COMANDO /start ===
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("üöÄ Benvenuto su ArcadiaAI!\nUsa /modello per scegliere il modello AI.")

# === COMANDO /reset ===
async def reset(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    user_histories[user_id] = []
    await update.message.reply_text("üß† Memoria contestuale resettata.")

# === COMANDO /modello ===
async def scegli_modello(update: Update, context: ContextTypes.DEFAULT_TYPE):
    keyboard = [
        [
            InlineKeyboardButton("CES 1.5", callback_data="gemini"),
            InlineKeyboardButton("CES Plus", callback_data="cesplus"),
        ],
        [
            InlineKeyboardButton("DeepSeek R1", callback_data="deepseek"),
        ]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_text("ü§ñ Scegli il modello AI da usare:", reply_markup=reply_markup)

# === CALLBACK: selezione modello ===
async def handle_model_selection(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    user_id = query.from_user.id
    modello = query.data.replace("modello_", "")

    nomi_modelli = {
        "gemini": "CES 1.5",
        "cesplus": "CES Plus",
        "deepseek": "DeepSeek R1"
    }

    user_models[user_id] = modello
    nome_modello = nomi_modelli.get(modello, "Modello sconosciuto")

    await query.edit_message_text(f"‚úÖ Hai selezionato il modello: *{nome_modello}*", parse_mode="Markdown")

# === AVVIO FLASK ===
def run_flask():
    port = int(os.environ.get("PORT", 5000))
    print(f"üåê Avvio server Flask sulla porta {port}...")
    app.run(host="0.0.0.0", port=port, debug=False, use_reloader=False)

# === AVVIO TELEGRAM BOT ===
async def run_telegram_bot():
    application = Application.builder().token(TELEGRAM_TOKEN).build()

    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("reset", reset))
    application.add_handler(CommandHandler("modello", scegli_modello))
    application.add_handler(CallbackQueryHandler(handle_model_selection))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, telegram_message_handler))

    print("ü§ñ Bot Telegram avviato!")
    await application.run_polling(
        allowed_updates=Update.ALL_TYPES,
        stop_signals=None,
        close_loop=False
    )

# === AVVIO PARALLELO ===
if __name__ == "__main__":
    flask_thread = threading.Thread(target=run_flask)
    flask_thread.start()

    asyncio.run(run_telegram_bot())
